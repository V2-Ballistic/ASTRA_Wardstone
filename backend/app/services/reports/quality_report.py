"""
ASTRA — Quality Assessment Report
====================================
File: backend/app/services/reports/quality_report.py   ← NEW

Content:
  - Overall quality score distribution (histogram)
  - Per-requirement quality details with pass/fail
  - Most common quality issues (from quality_checker)
  - Prohibited terms usage summary
  - TBD / TBR tracking
  - Improvement recommendations

Formats: xlsx (openpyxl), pdf (reportlab)
"""

import io
import re
from datetime import datetime
from collections import Counter, defaultdict

from sqlalchemy.orm import Session

from app.models import Requirement
from app.services.reports.base import ReportGenerator, ReportOutput

try:
    from app.services.quality_checker import check_requirement_quality
except ImportError:
    def check_requirement_quality(s, t="", r=""):
        return {"score": 0, "passed": False, "issues": [], "warnings": []}


# Prohibited terms from NASA Appendix C
PROHIBITED_TERMS = [
    "adequate", "appropriate", "as applicable", "as required",
    "be able to", "be capable of", "but not limited to", "capability of",
    "effective", "etc", "if practical", "normal", "provide for",
    "sufficient", "suitable", "timely", "TBD", "TBR",
]


class QualityReport(ReportGenerator):
    name = "quality"
    supported_formats = ["xlsx", "pdf"]

    def generate(self, project_id: int, db: Session, options: dict | None = None) -> ReportOutput:
        options = options or {}
        fmt = options.get("format", "xlsx")
        project = self._get_project(db, project_id)
        reqs = self._get_requirements(db, project_id)

        # Re-run quality checks for live data
        results = []
        scores = []
        all_issues: list[str] = []
        term_counts: Counter = Counter()
        tbd_reqs = []
        tbr_reqs = []

        for req in reqs:
            qr = check_requirement_quality(
                req.statement or "", req.title or "", req.rationale or ""
            )
            score = qr.get("score", req.quality_score or 0)
            passed = qr.get("passed", score >= 70)
            issues = qr.get("issues", []) + qr.get("warnings", [])
            scores.append(score)
            all_issues.extend(issues)

            # Prohibited terms scan
            text = f"{req.statement} {req.rationale or ''} {req.title}".lower()
            for term in PROHIBITED_TERMS:
                if term.lower() in text:
                    term_counts[term] += 1

            if "TBD" in (req.statement or "").upper():
                tbd_reqs.append(req.req_id)
            if "TBR" in (req.statement or "").upper():
                tbr_reqs.append(req.req_id)

            results.append({
                "req_id": req.req_id,
                "title": req.title,
                "type": self._enum_val(req.req_type),
                "level": self._enum_val(req.level) if hasattr(req, "level") else "L1",
                "score": round(score, 1),
                "passed": passed,
                "issues": "; ".join(issues) if issues else "—",
            })

        # Score distribution (histogram buckets)
        buckets = {"0-19": 0, "20-39": 0, "40-59": 0, "60-69": 0,
                   "70-79": 0, "80-89": 0, "90-100": 0}
        for s in scores:
            if s < 20: buckets["0-19"] += 1
            elif s < 40: buckets["20-39"] += 1
            elif s < 60: buckets["40-59"] += 1
            elif s < 70: buckets["60-69"] += 1
            elif s < 80: buckets["70-79"] += 1
            elif s < 90: buckets["80-89"] += 1
            else: buckets["90-100"] += 1

        # By type
        type_scores: dict[str, list] = defaultdict(list)
        for r, req in zip(results, reqs):
            type_scores[r["type"]].append(r["score"])
        by_type = {t: round(sum(s)/len(s), 1) if s else 0 for t, s in type_scores.items()}

        # Top issues
        issue_counter = Counter(all_issues)
        top_issues = issue_counter.most_common(10)

        avg = round(sum(scores) / len(scores), 1) if scores else 0
        passed_count = sum(1 for r in results if r["passed"])

        summary = {
            "total": len(reqs),
            "avg_score": avg,
            "passed": passed_count,
            "failed": len(reqs) - passed_count,
            "pass_rate": round(passed_count / len(reqs) * 100, 1) if reqs else 0,
            "tbd_count": len(tbd_reqs),
            "tbr_count": len(tbr_reqs),
            "generated_at": datetime.utcnow().isoformat(),
            "project": project.code,
        }

        if fmt == "xlsx":
            return self._to_xlsx(results, summary, buckets, by_type, top_issues,
                                 term_counts, tbd_reqs, tbr_reqs, project)
        return self._to_pdf(results, summary, buckets, by_type, top_issues,
                            term_counts, tbd_reqs, tbr_reqs, project)

    # ── Excel ────────────────────────────────────────────

    def _to_xlsx(self, results, summary, buckets, by_type, top_issues,
                 term_counts, tbd_reqs, tbr_reqs, project) -> ReportOutput:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.chart import BarChart, Reference

        wb = Workbook()

        # ── Sheet 1: Overview ──
        ws = wb.active
        ws.title = "Quality Overview"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1F2937")
        green_fill = PatternFill("solid", fgColor="D1FAE5")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        thin = Border(left=Side("thin"), right=Side("thin"),
                      top=Side("thin"), bottom=Side("thin"))

        ws.merge_cells("A1:G1")
        ws["A1"] = f"Quality Assessment Report — {project.code}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Generated: {summary['generated_at']}"

        # Summary block
        stats = [
            ("Total Requirements", summary["total"]),
            ("Average Quality Score", summary["avg_score"]),
            ("Passed (≥70)", summary["passed"]),
            ("Failed (<70)", summary["failed"]),
            ("Pass Rate", f"{summary['pass_rate']}%"),
            ("TBD Count", summary["tbd_count"]),
            ("TBR Count", summary["tbr_count"]),
        ]
        for i, (label, val) in enumerate(stats):
            ws.cell(row=4 + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=4 + i, column=2, value=val)

        # Score distribution
        ws.cell(row=13, column=1, value="Score Distribution").font = Font(bold=True, size=12)
        for i, (bucket, count) in enumerate(buckets.items()):
            ws.cell(row=14 + i, column=1, value=bucket)
            ws.cell(row=14 + i, column=2, value=count)

        # Bar chart
        chart = BarChart()
        chart.title = "Quality Score Distribution"
        chart.y_axis.title = "Requirements"
        data = Reference(ws, min_col=2, min_row=14, max_row=14 + len(buckets) - 1)
        cats = Reference(ws, min_col=1, min_row=14, max_row=14 + len(buckets) - 1)
        chart.add_data(data)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "D4")

        # By type
        row = 23
        ws.cell(row=row, column=1, value="Average Score by Type").font = Font(bold=True, size=12)
        for i, (t, avg) in enumerate(sorted(by_type.items())):
            ws.cell(row=row + 1 + i, column=1, value=t.title())
            ws.cell(row=row + 1 + i, column=2, value=avg)

        # Column widths
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 14

        # ── Sheet 2: Per-requirement details ──
        ws2 = wb.create_sheet("Requirements Detail")
        headers = ["Req ID", "Title", "Type", "Level", "Score", "Passed", "Issues"]
        for col, h in enumerate(headers, 1):
            c = ws2.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.border = thin

        for i, r in enumerate(results, 2):
            ws2.cell(row=i, column=1, value=r["req_id"]).border = thin
            ws2.cell(row=i, column=2, value=r["title"]).border = thin
            ws2.cell(row=i, column=3, value=r["type"]).border = thin
            ws2.cell(row=i, column=4, value=r["level"]).border = thin
            sc = ws2.cell(row=i, column=5, value=r["score"])
            sc.border = thin
            pc = ws2.cell(row=i, column=6, value="PASS" if r["passed"] else "FAIL")
            pc.border = thin
            pc.fill = green_fill if r["passed"] else red_fill
            ws2.cell(row=i, column=7, value=r["issues"]).border = thin

        for col, w in zip("ABCDEFG", [14, 40, 14, 8, 8, 8, 50]):
            ws2.column_dimensions[col].width = w

        # ── Sheet 3: Issues & Terms ──
        ws3 = wb.create_sheet("Issues & Terms")
        ws3["A1"] = "Top Quality Issues"
        ws3["A1"].font = Font(bold=True, size=12)
        for i, (issue, count) in enumerate(top_issues):
            ws3.cell(row=2 + i, column=1, value=issue)
            ws3.cell(row=2 + i, column=2, value=count)

        row = 3 + len(top_issues)
        ws3.cell(row=row, column=1, value="Prohibited Terms Found").font = Font(bold=True, size=12)
        for i, (term, count) in enumerate(term_counts.most_common(20)):
            ws3.cell(row=row + 1 + i, column=1, value=term)
            ws3.cell(row=row + 1 + i, column=2, value=count)

        row2 = row + 2 + len(term_counts)
        if tbd_reqs:
            ws3.cell(row=row2, column=1, value="TBD Requirements").font = Font(bold=True, size=12)
            for i, rid in enumerate(tbd_reqs):
                ws3.cell(row=row2 + 1 + i, column=1, value=rid)
        if tbr_reqs:
            row3 = row2 + 2 + len(tbd_reqs)
            ws3.cell(row=row3, column=1, value="TBR Requirements").font = Font(bold=True, size=12)
            for i, rid in enumerate(tbr_reqs):
                ws3.cell(row=row3 + 1 + i, column=1, value=rid)

        ws3.column_dimensions["A"].width = 50
        ws3.column_dimensions["B"].width = 10

        buf = io.BytesIO()
        wb.save(buf)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"Quality_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata=summary,
        )

    # ── PDF ──────────────────────────────────────────────

    def _to_pdf(self, results, summary, buckets, by_type, top_issues,
                term_counts, tbd_reqs, tbr_reqs, project) -> ReportOutput:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
        )
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=LETTER)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            f"<b>Quality Assessment Report — {project.code}</b>", styles["Title"]))
        elements.append(Paragraph(f"Generated: {summary['generated_at']}", styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Summary
        elements.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
        sum_data = [
            ["Metric", "Value"],
            ["Total Requirements", str(summary["total"])],
            ["Average Score", str(summary["avg_score"])],
            ["Pass Rate", f"{summary['pass_rate']}%"],
            ["TBD Items", str(summary["tbd_count"])],
            ["TBR Items", str(summary["tbr_count"])],
        ]
        t = Table(sum_data, colWidths=[3*inch, 2*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 16))

        # Score distribution table
        elements.append(Paragraph("<b>Score Distribution</b>", styles["Heading2"]))
        dist_data = [["Range", "Count"]] + [[k, str(v)] for k, v in buckets.items()]
        dt = Table(dist_data, colWidths=[2*inch, 1.5*inch])
        dt.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(dt)
        elements.append(Spacer(1, 16))

        # Top issues
        if top_issues:
            elements.append(Paragraph("<b>Top Quality Issues</b>", styles["Heading2"]))
            for issue, count in top_issues[:8]:
                elements.append(Paragraph(
                    f"• ({count}x) {issue}", styles["Normal"]))
            elements.append(Spacer(1, 12))

        # Requirements table
        elements.append(PageBreak())
        elements.append(Paragraph("<b>Per-Requirement Quality</b>", styles["Heading2"]))
        header = ["ID", "Title", "Type", "Score", "Pass"]
        data = [header]
        row_colors = []
        for r in results:
            data.append([
                r["req_id"], r["title"][:35], r["type"],
                str(r["score"]), "PASS" if r["passed"] else "FAIL",
            ])
            row_colors.append(
                colors.Color(0.82, 0.98, 0.90) if r["passed"]
                else colors.Color(1, 0.89, 0.89)
            )

        rt = Table(data, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
        ]
        for i, c in enumerate(row_colors):
            style_cmds.append(("BACKGROUND", (4, i + 1), (4, i + 1), c))
        rt.setStyle(TableStyle(style_cmds))
        elements.append(rt)

        doc.build(elements)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"Quality_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
            content_type="application/pdf",
            metadata=summary,
        )

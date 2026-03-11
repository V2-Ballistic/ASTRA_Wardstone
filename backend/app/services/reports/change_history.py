"""
ASTRA — Change History Report
================================
File: backend/app/services/reports/change_history.py   ← NEW

Detailed change history for CCB (Configuration Control Board) use:
  - All changes within a date range
  - Grouped by requirement, showing field diffs
  - Who changed what, when

Formats: xlsx (openpyxl), pdf (reportlab)
"""

import io
from datetime import datetime
from collections import defaultdict

from sqlalchemy.orm import Session

from app.models import Requirement, RequirementHistory, User
from app.services.reports.base import ReportGenerator, ReportOutput


class ChangeHistoryReport(ReportGenerator):
    name = "change-history"
    supported_formats = ["xlsx", "pdf"]

    def generate(self, project_id: int, db: Session, options: dict | None = None) -> ReportOutput:
        options = options or {}
        fmt = options.get("format", "xlsx")
        date_from = options.get("date_from")
        date_to = options.get("date_to")

        project = self._get_project(db, project_id)
        reqs = self._get_requirements(db, project_id)
        req_ids = [r.id for r in reqs]
        req_map = {r.id: r for r in reqs}

        history = self._get_history(db, req_ids, date_from, date_to)

        # Group by requirement
        grouped: dict[str, list] = defaultdict(list)
        flat_rows = []
        for h in history:
            req = req_map.get(h.requirement_id)
            user = db.query(User).filter(User.id == h.changed_by_id).first() if h.changed_by_id else None
            entry = {
                "req_id": req.req_id if req else "—",
                "title": req.title if req else "—",
                "field": h.field_changed or "—",
                "old_value": (h.old_value or "")[:80],
                "new_value": (h.new_value or "")[:80],
                "description": h.change_description or "—",
                "changed_by": user.full_name if user else "System",
                "changed_at": self._ts(h.changed_at),
                "version": h.version,
            }
            grouped[entry["req_id"]].append(entry)
            flat_rows.append(entry)

        summary = {
            "project": project.code,
            "total_changes": len(flat_rows),
            "requirements_affected": len(grouped),
            "date_from": date_from.isoformat() if date_from else "—",
            "date_to": date_to.isoformat() if date_to else "—",
            "generated_at": datetime.utcnow().isoformat(),
        }

        if fmt == "xlsx":
            return self._to_xlsx(flat_rows, grouped, summary, project)
        return self._to_pdf(flat_rows, grouped, summary, project)

    def _to_xlsx(self, flat_rows, grouped, summary, project) -> ReportOutput:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Change History"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1F2937")
        thin = Border(left=Side("thin"), right=Side("thin"),
                      top=Side("thin"), bottom=Side("thin"))

        ws.merge_cells("A1:H1")
        ws["A1"] = f"Change History — {project.code}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (f"Period: {summary['date_from']} to {summary['date_to']} | "
                     f"Changes: {summary['total_changes']} | "
                     f"Requirements affected: {summary['requirements_affected']}")

        headers = ["Date", "Req ID", "Title", "Field", "Old Value", "New Value", "Changed By", "Version"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.border = thin

        for i, row in enumerate(flat_rows, 5):
            vals = [
                row["changed_at"], row["req_id"], row["title"][:40],
                row["field"], row["old_value"], row["new_value"],
                row["changed_by"], row["version"],
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.border = thin

        widths = [18, 14, 30, 16, 25, 25, 18, 8]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Grouped sheet
        ws2 = wb.create_sheet("By Requirement")
        row_num = 1
        for req_id in sorted(grouped.keys()):
            changes = grouped[req_id]
            ws2.cell(row=row_num, column=1, value=req_id).font = Font(bold=True, size=12)
            ws2.cell(row=row_num, column=2, value=changes[0]["title"] if changes else "")
            row_num += 1
            for ch in changes:
                ws2.cell(row=row_num, column=2, value=ch["changed_at"])
                ws2.cell(row=row_num, column=3, value=ch["field"])
                ws2.cell(row=row_num, column=4, value=ch["old_value"])
                ws2.cell(row=row_num, column=5, value=ch["new_value"])
                ws2.cell(row=row_num, column=6, value=ch["changed_by"])
                row_num += 1
            row_num += 1  # blank row between requirements

        for col, w in zip("ABCDEF", [14, 18, 16, 25, 25, 18]):
            ws2.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"ChangeHistory_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata=summary,
        )

    def _to_pdf(self, flat_rows, grouped, summary, project) -> ReportOutput:
        from reportlab.lib.pagesizes import landscape, LETTER
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(LETTER),
                                leftMargin=0.5*inch, rightMargin=0.5*inch)
        styles = getSampleStyleSheet()
        el = []

        el.append(Paragraph(
            f"<b>Change History — {project.code}</b>", styles["Title"]))
        el.append(Paragraph(
            f"Period: {summary['date_from']} → {summary['date_to']} | "
            f"Total changes: {summary['total_changes']} across "
            f"{summary['requirements_affected']} requirements",
            styles["Normal"]))
        el.append(Spacer(1, 16))

        dark = colors.Color(0.12, 0.16, 0.22)
        grid_c = colors.Color(0.8, 0.8, 0.8)

        # Grouped by requirement
        for req_id in sorted(grouped.keys()):
            changes = grouped[req_id]
            el.append(Paragraph(
                f"<b>{req_id}</b> — {changes[0]['title'][:60]}", styles["Heading3"]))

            data = [["Date", "Field", "Old Value", "New Value", "By"]]
            for ch in changes:
                data.append([
                    ch["changed_at"], ch["field"],
                    ch["old_value"][:40], ch["new_value"][:40],
                    ch["changed_by"],
                ])

            t = Table(data, repeatRows=1,
                      colWidths=[1.3*inch, 1.2*inch, 2.5*inch, 2.5*inch, 1.3*inch])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), dark),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, grid_c),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            el.append(t)
            el.append(Spacer(1, 12))

        doc.build(el)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"ChangeHistory_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
            content_type="application/pdf",
            metadata=summary,
        )

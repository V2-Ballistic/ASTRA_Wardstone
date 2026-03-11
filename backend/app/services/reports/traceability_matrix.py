"""
ASTRA — Traceability Matrix Report
====================================
File: backend/app/services/reports/traceability_matrix.py   ← NEW

Generates a full Requirements Traceability Matrix (RTM) with:
  - Req ID, title, status, priority, level, parent
  - Source artifacts linked to each requirement
  - Child decomposition count
  - Verification methods and verification status
  - Color coding: green = fully traced, red = orphan, yellow = partial

Formats: xlsx (openpyxl), pdf (reportlab), html
"""

import io
import json
from datetime import datetime

from sqlalchemy.orm import Session

from app.models import Requirement, SourceArtifact
from app.services.reports.base import ReportGenerator, ReportOutput


class TraceabilityMatrixReport(ReportGenerator):
    name = "traceability-matrix"
    supported_formats = ["xlsx", "pdf", "html"]

    def generate(self, project_id: int, db: Session, options: dict | None = None) -> ReportOutput:
        options = options or {}
        fmt = options.get("format", "xlsx")
        project = self._get_project(db, project_id)
        reqs = self._get_requirements(db, project_id)
        req_ids = [r.id for r in reqs]
        links = self._get_trace_links(db, req_ids)
        verifs = self._get_verifications(db, req_ids)

        # Build per-requirement trace data
        rows = []
        orphan_count = 0
        partial_count = 0
        full_count = 0

        for req in reqs:
            req_links = [l for l in links if l.source_id == req.id or l.target_id == req.id]
            src_arts = [l.source_id for l in req_links if l.source_type == "source_artifact"]
            children = [l.target_id for l in req_links if l.link_type and
                        (self._enum_val(l.link_type) == "decomposition") and l.source_id == req.id]
            v_list = verifs.get(req.id, [])
            v_methods = ", ".join(set(self._enum_val(v.method) for v in v_list)) or "None"
            v_statuses = ", ".join(set(self._enum_val(v.status) for v in v_list)) or "N/A"

            has_source = len(src_arts) > 0
            has_verif = len(v_list) > 0
            if has_source and has_verif:
                trace_status = "Full"
                full_count += 1
            elif has_source or has_verif or len(children) > 0:
                trace_status = "Partial"
                partial_count += 1
            else:
                trace_status = "Orphan"
                orphan_count += 1

            rows.append({
                "req_id": req.req_id,
                "title": req.title,
                "status": self._enum_val(req.status),
                "priority": self._enum_val(req.priority),
                "level": self._enum_val(req.level) if hasattr(req, "level") else "L1",
                "parent": req.parent.req_id if req.parent else "—",
                "source_artifacts": len(src_arts),
                "children": len(children),
                "verification_methods": v_methods,
                "verification_status": v_statuses,
                "trace_status": trace_status,
            })

        summary = {
            "total": len(reqs),
            "fully_traced": full_count,
            "partial": partial_count,
            "orphans": orphan_count,
            "generated_at": datetime.utcnow().isoformat(),
            "project": project.code,
        }

        if fmt == "xlsx":
            return self._to_xlsx(rows, summary, project)
        elif fmt == "html":
            return self._to_html(rows, summary, project)
        else:
            return self._to_pdf(rows, summary, project)

    # ── Excel ────────────────────────────────────────────

    def _to_xlsx(self, rows, summary, project) -> ReportOutput:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Traceability Matrix"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1F2937")
        green_fill = PatternFill("solid", fgColor="D1FAE5")
        yellow_fill = PatternFill("solid", fgColor="FEF3C7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Title row
        ws.merge_cells("A1:K1")
        ws["A1"] = f"Requirements Traceability Matrix — {project.code} {project.name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A2:K2")
        ws["A2"] = f"Generated: {summary['generated_at']}"
        ws["A2"].font = Font(italic=True, color="6B7280")

        # Headers
        headers = [
            "Req ID", "Title", "Status", "Priority", "Level",
            "Parent", "Source Artifacts", "Children",
            "Verification Methods", "Verification Status", "Trace Status",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        for i, row in enumerate(rows, 5):
            vals = list(row.values())
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=i, column=col, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Color the Trace Status column
            status_cell = ws.cell(row=i, column=11)
            if row["trace_status"] == "Full":
                status_cell.fill = green_fill
            elif row["trace_status"] == "Partial":
                status_cell.fill = yellow_fill
            else:
                status_cell.fill = red_fill

        # Summary section
        sum_row = len(rows) + 6
        ws.cell(row=sum_row, column=1, value="Summary").font = Font(bold=True, size=12)
        for j, (label, val) in enumerate([
            ("Total Requirements", summary["total"]),
            ("Fully Traced", summary["fully_traced"]),
            ("Partially Traced", summary["partial"]),
            ("Orphans (no links)", summary["orphans"]),
        ]):
            ws.cell(row=sum_row + 1 + j, column=1, value=label)
            ws.cell(row=sum_row + 1 + j, column=2, value=val)

        # Column widths
        widths = [14, 40, 12, 10, 8, 14, 16, 10, 24, 20, 14]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"RTM_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata=summary,
        )

    # ── HTML ─────────────────────────────────────────────

    def _to_html(self, rows, summary, project) -> ReportOutput:
        color_map = {"Full": "#D1FAE5", "Partial": "#FEF3C7", "Orphan": "#FEE2E2"}
        table_rows = ""
        for r in rows:
            bg = color_map.get(r["trace_status"], "#FFF")
            table_rows += f"""<tr>
                <td>{r['req_id']}</td><td>{r['title']}</td><td>{r['status']}</td>
                <td>{r['priority']}</td><td>{r['level']}</td><td>{r['parent']}</td>
                <td>{r['source_artifacts']}</td><td>{r['children']}</td>
                <td>{r['verification_methods']}</td><td>{r['verification_status']}</td>
                <td style="background:{bg};font-weight:bold">{r['trace_status']}</td>
            </tr>"""

        html = f"""<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8">
<title>RTM — {project.code}</title>
<style>
  body {{ font-family: 'Segoe UI', sans-serif; margin: 2rem; color: #1F2937; }}
  h1 {{ font-size: 1.5rem; }} h2 {{ font-size: 1.1rem; margin-top: 2rem; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
  th {{ background: #1F2937; color: #FFF; padding: 8px; text-align: left; }}
  td {{ padding: 6px 8px; border: 1px solid #E5E7EB; vertical-align: top; }}
  tr:nth-child(even) {{ background: #F9FAFB; }}
  .summary {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 1rem; margin-top: 1rem; }}
  .stat {{ padding: 1rem; border-radius: 8px; background: #F3F4F6; text-align: center; }}
  .stat strong {{ font-size: 1.5rem; display: block; }}
</style></head><body>
<h1>Requirements Traceability Matrix — {project.code}</h1>
<p>Generated: {summary['generated_at']}</p>
<table><thead><tr>
  <th>Req ID</th><th>Title</th><th>Status</th><th>Priority</th><th>Level</th>
  <th>Parent</th><th>Source Arts</th><th>Children</th>
  <th>Verif Methods</th><th>Verif Status</th><th>Trace Status</th>
</tr></thead><tbody>{table_rows}</tbody></table>
<h2>Summary</h2>
<div class="summary">
  <div class="stat"><strong>{summary['total']}</strong>Total</div>
  <div class="stat" style="background:#D1FAE5"><strong>{summary['fully_traced']}</strong>Fully Traced</div>
  <div class="stat" style="background:#FEF3C7"><strong>{summary['partial']}</strong>Partial</div>
  <div class="stat" style="background:#FEE2E2"><strong>{summary['orphans']}</strong>Orphans</div>
</div></body></html>"""

        return ReportOutput(
            content=html.encode(),
            filename=f"RTM_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.html",
            content_type="text/html",
            metadata=summary,
        )

    # ── PDF ──────────────────────────────────────────────

    def _to_pdf(self, rows, summary, project) -> ReportOutput:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, LETTER
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(LETTER),
                                leftMargin=0.5*inch, rightMargin=0.5*inch)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            f"<b>Requirements Traceability Matrix — {project.code}</b>",
            styles["Title"]))
        elements.append(Paragraph(f"Generated: {summary['generated_at']}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        # Table
        header = ["ID", "Title", "Status", "Pri", "Lvl", "Parent",
                   "Src", "Ch", "V.Meth", "V.Status", "Trace"]
        data = [header]
        trace_colors = []
        for i, r in enumerate(rows):
            data.append([
                r["req_id"], r["title"][:40], r["status"], r["priority"],
                r["level"], r["parent"], str(r["source_artifacts"]),
                str(r["children"]), r["verification_methods"][:20],
                r["verification_status"][:15], r["trace_status"],
            ])
            c = (colors.Color(0.82, 0.98, 0.90) if r["trace_status"] == "Full"
                 else colors.Color(1, 0.96, 0.78) if r["trace_status"] == "Partial"
                 else colors.Color(1, 0.89, 0.89))
            trace_colors.append((i + 1, c))

        t = Table(data, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for row_idx, bg_color in trace_colors:
            style_cmds.append(("BACKGROUND", (10, row_idx), (10, row_idx), bg_color))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        # Summary
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
        elements.append(Paragraph(
            f"Total: {summary['total']}  |  Fully Traced: {summary['fully_traced']}  |  "
            f"Partial: {summary['partial']}  |  Orphans: {summary['orphans']}",
            styles["Normal"]))

        doc.build(elements)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"RTM_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
            content_type="application/pdf",
            metadata=summary,
        )

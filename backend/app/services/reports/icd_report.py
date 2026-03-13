"""
ASTRA — Interface Control Document (ICD) Report
===================================================
File: backend/app/services/reports/icd_report.py
Path: C:\\Users\\Mason\\Documents\\ASTRA\\backend\\app\\services\\reports\\icd_report.py

Generates a comprehensive ICD as .xlsx with 9 sheets:
  1. Interface Summary — N² matrix of unit connections
  2. Unit Catalog — all units with key specs
  3. Connector Pinouts — all pin tables per connector
  4. Bus Configuration — bus definitions with pin assignments
  5. Message Catalog — all messages with field details
  6. Wire Harness List — harness wire tables
  7. Signal Dictionary — complete signal cross-reference
  8. Environmental Summary — unit env/EMI specs + requirement status
  9. Requirements Trace — interface → requirement links with status

Registered as "icd" in the REPORT_REGISTRY.
"""

import io
from datetime import datetime
from collections import defaultdict

from sqlalchemy.orm import Session
from sqlalchemy import func

from app.services.reports.base import ReportGenerator, ReportOutput

# Interface models — graceful import
try:
    from app.models.interface import (
        System, Unit, Connector, Pin, BusDefinition,
        PinBusAssignment, MessageDefinition, MessageField,
        WireHarness, Wire, Interface,
        UnitEnvironmentalSpec, InterfaceRequirementLink,
    )
    _HAS_INTERFACE = True
except ImportError:
    _HAS_INTERFACE = False


class ICDReport(ReportGenerator):
    name = "icd"
    supported_formats = ["xlsx"]

    def generate(self, project_id: int, db: Session, options: dict | None = None) -> ReportOutput:
        if not _HAS_INTERFACE:
            raise ValueError("Interface module not installed")

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        project = self._get_project(db, project_id)
        wb = Workbook()

        # ── Shared styles ──
        hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill("solid", fgColor="1F2937")
        title_font = Font(name="Arial", bold=True, size=14)
        subtitle_font = Font(name="Arial", bold=True, size=11)
        body_font = Font(name="Arial", size=9)
        mono_font = Font(name="Consolas", size=9)
        thin = Border(
            left=Side("thin", color="D1D5DB"),
            right=Side("thin", color="D1D5DB"),
            top=Side("thin", color="D1D5DB"),
            bottom=Side("thin", color="D1D5DB"),
        )
        center = Alignment(horizontal="center", vertical="center")
        wrap = Alignment(wrap_text=True, vertical="top")
        blue_fill = PatternFill("solid", fgColor="DBEAFE")
        green_fill = PatternFill("solid", fgColor="D1FAE5")
        yellow_fill = PatternFill("solid", fgColor="FEF3C7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        gray_fill = PatternFill("solid", fgColor="F3F4F6")

        ev = self._enum_val

        # ── Load all data once ──
        systems = db.query(System).filter(System.project_id == project_id).order_by(System.name).all()
        units = db.query(Unit).filter(Unit.project_id == project_id).order_by(Unit.designation).all()
        connectors = db.query(Connector).filter(Connector.project_id == project_id).order_by(Connector.designator).all()
        buses = db.query(BusDefinition).filter(BusDefinition.project_id == project_id).order_by(BusDefinition.name).all()
        harnesses = db.query(WireHarness).filter(WireHarness.project_id == project_id).all()
        interfaces = db.query(Interface).filter(Interface.project_id == project_id).all()
        links = db.query(InterfaceRequirementLink).all()
        env_specs = db.query(UnitEnvironmentalSpec).join(Unit).filter(Unit.project_id == project_id).all()

        unit_map = {u.id: u for u in units}
        sys_map = {s.id: s for s in systems}
        conn_map = {c.id: c for c in connectors}

        def _write_header(ws, row, headers, widths=None):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col, value=h)
                c.font = hdr_font
                c.fill = hdr_fill
                c.border = thin
                c.alignment = center
            if widths:
                for col, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(col)].width = w

        def _write_row(ws, row, values, font=body_font, fills=None):
            for col, v in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = font
                c.border = thin
                c.alignment = Alignment(vertical="top")
                if fills and col in fills:
                    c.fill = fills[col]

        # ══════════════════════════════════════
        #  Sheet 1: Interface Summary (N² matrix)
        # ══════════════════════════════════════

        ws1 = wb.active
        ws1.title = "Interface Summary"
        ws1.merge_cells("A1:F1")
        ws1["A1"] = f"Interface Control Document — {project.code} {project.name}"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws1["A2"].font = Font(name="Arial", italic=True, size=9, color="6B7280")
        ws1["A4"] = "N² Interface Matrix"
        ws1["A4"].font = subtitle_font

        if systems:
            n = len(systems)
            # Header row
            for ci, s in enumerate(systems):
                c = ws1.cell(row=6, column=ci + 2, value=s.abbreviation or s.name[:10])
                c.font = hdr_font
                c.fill = hdr_fill
                c.border = thin
                c.alignment = center
                ws1.column_dimensions[get_column_letter(ci + 2)].width = 14
            ws1.column_dimensions["A"].width = 14

            # Row labels + data
            for ri, src in enumerate(systems):
                rc = ws1.cell(row=7 + ri, column=1, value=src.abbreviation or src.name[:10])
                rc.font = Font(name="Arial", bold=True, size=9)
                rc.fill = hdr_fill
                rc.font = hdr_font
                rc.border = thin

                for ci, tgt in enumerate(systems):
                    cell = ws1.cell(row=7 + ri, column=ci + 2)
                    cell.border = thin
                    cell.alignment = center
                    if ri == ci:
                        cell.value = "—"
                        cell.fill = gray_fill
                    else:
                        # Count interfaces + harnesses between these systems
                        icount = sum(
                            1 for iface in interfaces
                            if (iface.source_system_id == src.id and iface.target_system_id == tgt.id)
                            or (iface.source_system_id == tgt.id and iface.target_system_id == src.id)
                        )
                        # Count harnesses between units of these systems
                        src_unit_ids = {u.id for u in units if u.system_id == src.id}
                        tgt_unit_ids = {u.id for u in units if u.system_id == tgt.id}
                        hcount = sum(
                            1 for h in harnesses
                            if (h.from_unit_id in src_unit_ids and h.to_unit_id in tgt_unit_ids)
                            or (h.from_unit_id in tgt_unit_ids and h.to_unit_id in src_unit_ids)
                        )
                        if icount or hcount:
                            cell.value = f"{icount}I / {hcount}H"
                            cell.fill = blue_fill
                        else:
                            cell.value = ""

            # Legend
            legend_row = 7 + n + 2
            ws1.cell(row=legend_row, column=1, value="Legend:").font = Font(name="Arial", bold=True, size=9)
            ws1.cell(row=legend_row, column=2, value="I = Interfaces, H = Harnesses").font = body_font

        # Summary stats below matrix
        stats_row = (7 + len(systems) + 5) if systems else 6
        ws1.cell(row=stats_row, column=1, value="Summary").font = subtitle_font
        stats = [
            ("Systems", len(systems)),
            ("Units", len(units)),
            ("Connectors", len(connectors)),
            ("Bus Definitions", len(buses)),
            ("Wire Harnesses", len(harnesses)),
            ("Interfaces", len(interfaces)),
        ]
        for i, (label, val) in enumerate(stats):
            ws1.cell(row=stats_row + 1 + i, column=1, value=label).font = Font(name="Arial", bold=True, size=9)
            ws1.cell(row=stats_row + 1 + i, column=2, value=val).font = body_font

        # ══════════════════════════════════════
        #  Sheet 2: Unit Catalog
        # ══════════════════════════════════════

        ws2 = wb.create_sheet("Unit Catalog")
        ws2["A1"] = "Unit Catalog"
        ws2["A1"].font = title_font

        headers2 = [
            "Designation", "Name", "System", "Type", "Status", "P/N",
            "Manufacturer", "Mass (kg)", "Power (W)", "Voltage",
            "Temp Op (°C)", "Vib (Grms)", "Connectors", "Buses",
        ]
        widths2 = [14, 24, 16, 14, 14, 16, 16, 10, 10, 12, 16, 10, 10, 8]
        _write_header(ws2, 3, headers2, widths2)

        for i, u in enumerate(units, 4):
            sys = sys_map.get(u.system_id)
            conn_count = db.query(func.count(Connector.id)).filter(Connector.unit_id == u.id).scalar()
            bus_count = db.query(func.count(BusDefinition.id)).filter(BusDefinition.unit_id == u.id).scalar()
            temp_range = ""
            if u.temp_operating_min_c is not None and u.temp_operating_max_c is not None:
                temp_range = f"{u.temp_operating_min_c} to {u.temp_operating_max_c}"
            _write_row(ws2, i, [
                u.designation, u.name, sys.name if sys else "—",
                ev(u.unit_type).replace("_", " "), ev(u.status).replace("_", " "),
                u.part_number, u.manufacturer,
                u.mass_kg, u.power_watts_nominal, u.voltage_input_nominal,
                temp_range, u.vibration_random_grms,
                conn_count, bus_count,
            ])

        # ══════════════════════════════════════
        #  Sheet 3: Connector Pinouts
        # ══════════════════════════════════════

        ws3 = wb.create_sheet("Connector Pinouts")
        ws3["A1"] = "Connector Pinouts"
        ws3["A1"].font = title_font

        row3 = 3
        for conn in connectors:
            unit = unit_map.get(conn.unit_id)
            ws3.cell(row=row3, column=1,
                     value=f"{unit.designation if unit else '?'} / {conn.designator} — {conn.name or ev(conn.connector_type)}")
            ws3.cell(row=row3, column=1).font = subtitle_font
            ws3.cell(row=row3, column=5,
                     value=f"{ev(conn.connector_type)} {ev(conn.gender)} {conn.total_contacts} contacts")
            ws3.cell(row=row3, column=5).font = Font(name="Arial", italic=True, size=9, color="6B7280")
            row3 += 1

            pin_headers = ["Pin", "Label", "Signal Name", "Signal Type", "Direction",
                           "Voltage", "Current (A)", "Impedance (Ω)", "Bus Assignment"]
            pin_widths = [6, 12, 20, 18, 14, 10, 10, 12, 18]
            _write_header(ws3, row3, pin_headers, pin_widths)
            row3 += 1

            pins = db.query(Pin).filter(Pin.connector_id == conn.id).order_by(Pin.pin_number).all()
            for pin in pins:
                # Check bus assignment
                ba = db.query(PinBusAssignment).filter(PinBusAssignment.pin_id == pin.id).first()
                bus_info = ""
                if ba:
                    bus = db.query(BusDefinition).filter(BusDefinition.id == ba.bus_def_id).first()
                    bus_info = f"{bus.name} ({ev(ba.pin_role)})" if bus else ev(ba.pin_role)

                sig_type = ev(pin.signal_type)
                fills = {}
                if sig_type.startswith("power"):
                    fills = {4: red_fill}
                elif "spare" in sig_type or "no_connect" in sig_type:
                    fills = {4: gray_fill}
                elif sig_type.startswith("shield"):
                    fills = {4: yellow_fill}

                _write_row(ws3, row3, [
                    pin.pin_number, pin.pin_label or "", pin.signal_name,
                    sig_type.replace("_", " "), ev(pin.direction).replace("_", " "),
                    pin.voltage_nominal or "", pin.current_max_amps or "",
                    pin.impedance_ohms or "", bus_info,
                ], fills=fills)
                row3 += 1

            row3 += 1  # blank row between connectors

        # ══════════════════════════════════════
        #  Sheet 4: Bus Configuration
        # ══════════════════════════════════════

        ws4 = wb.create_sheet("Bus Configuration")
        ws4["A1"] = "Bus Configuration"
        ws4["A1"].font = title_font

        row4 = 3
        for bus in buses:
            unit = unit_map.get(bus.unit_id)
            ws4.cell(row=row4, column=1,
                     value=f"{bus.name} — {ev(bus.protocol).replace('_', '-').upper()}")
            ws4.cell(row=row4, column=1).font = subtitle_font
            row4 += 1

            # Bus details
            details = [
                ("Unit", unit.designation if unit else "—"),
                ("Role", ev(bus.bus_role).replace("_", " ")),
                ("Address", bus.bus_address or "—"),
                ("Data Rate", bus.data_rate or "—"),
                ("Network Name", bus.bus_name_network or "—"),
                ("Word Size", f"{bus.word_size_bits} bits" if bus.word_size_bits else "—"),
                ("Topology", ev(bus.topology) if bus.topology else "—"),
                ("Redundancy", ev(bus.redundancy) if bus.redundancy else "—"),
            ]
            for label, val in details:
                ws4.cell(row=row4, column=1, value=label).font = Font(name="Arial", bold=True, size=9)
                ws4.cell(row=row4, column=2, value=val).font = body_font
                row4 += 1

            # Pin assignments
            pas = db.query(PinBusAssignment).filter(PinBusAssignment.bus_def_id == bus.id).all()
            if pas:
                row4 += 1
                ws4.cell(row=row4, column=1, value="Pin Assignments:").font = Font(name="Arial", bold=True, size=9)
                row4 += 1
                _write_header(ws4, row4, ["Pin", "Signal", "Connector", "Role"], [8, 18, 14, 14])
                row4 += 1
                for pa in pas:
                    pin = db.query(Pin).filter(Pin.id == pa.pin_id).first()
                    conn = conn_map.get(pin.connector_id) if pin else None
                    _write_row(ws4, row4, [
                        pin.pin_number if pin else "?",
                        pin.signal_name if pin else "?",
                        conn.designator if conn else "?",
                        ev(pa.pin_role).replace("_", " "),
                    ])
                    row4 += 1

            row4 += 2

        # ══════════════════════════════════════
        #  Sheet 5: Message Catalog
        # ══════════════════════════════════════

        ws5 = wb.create_sheet("Message Catalog")
        ws5["A1"] = "Message Catalog"
        ws5["A1"].font = title_font

        row5 = 3
        for bus in buses:
            msgs = db.query(MessageDefinition).filter(
                MessageDefinition.bus_def_id == bus.id
            ).order_by(MessageDefinition.label).all()
            if not msgs:
                continue

            ws5.cell(row=row5, column=1,
                     value=f"Bus: {bus.name} ({ev(bus.protocol).replace('_', '-').upper()})")
            ws5.cell(row=row5, column=1).font = subtitle_font
            row5 += 1

            for msg in msgs:
                unit = unit_map.get(msg.unit_id)
                ws5.cell(row=row5, column=1,
                         value=f"{msg.label} ({msg.mnemonic or '—'})")
                ws5.cell(row=row5, column=1).font = Font(name="Arial", bold=True, size=9)
                ws5.cell(row=row5, column=3,
                         value=f"Dir: {ev(msg.direction)} | Rate: {msg.rate_hz or '—'} Hz | Words: {msg.word_count or '—'} | SA: {msg.subaddress or '—'}")
                ws5.cell(row=row5, column=3).font = Font(name="Arial", italic=True, size=8, color="6B7280")
                row5 += 1

                # Fields
                fields = db.query(MessageField).filter(
                    MessageField.message_id == msg.id
                ).order_by(MessageField.field_order, MessageField.word_number).all()

                if fields:
                    fld_headers = ["Field Name", "Label", "Type", "Word", "Bit Offset",
                                   "Bit Length", "Min", "Max", "Unit", "Scale"]
                    fld_widths = [16, 16, 12, 6, 8, 8, 10, 10, 10, 8]
                    _write_header(ws5, row5, fld_headers, fld_widths)
                    row5 += 1

                    for fld in fields:
                        fills = {}
                        if fld.is_spare or fld.is_padding:
                            fills = {1: gray_fill, 2: gray_fill}
                        _write_row(ws5, row5, [
                            fld.field_name, fld.label or "",
                            ev(fld.data_type).replace("_", " "),
                            fld.word_number, fld.bit_offset, fld.bit_length,
                            fld.min_value, fld.max_value,
                            fld.unit_of_measure or "", fld.scale_factor,
                        ], font=mono_font, fills=fills)
                        row5 += 1

                row5 += 1
            row5 += 1

        # ══════════════════════════════════════
        #  Sheet 6: Wire Harness List
        # ══════════════════════════════════════

        ws6 = wb.create_sheet("Wire Harness List")
        ws6["A1"] = "Wire Harness List"
        ws6["A1"].font = title_font

        row6 = 3
        for harness in harnesses:
            fu = unit_map.get(harness.from_unit_id)
            tu = unit_map.get(harness.to_unit_id)
            fc = conn_map.get(harness.from_connector_id)
            tc = conn_map.get(harness.to_connector_id)

            ws6.cell(row=row6, column=1,
                     value=f"{harness.harness_id or f'HAR-{harness.id}'} — {harness.name}")
            ws6.cell(row=row6, column=1).font = subtitle_font
            row6 += 1
            ws6.cell(row=row6, column=1,
                     value=f"From: {fu.designation if fu else '?'} ({fc.designator if fc else '?'}) → To: {tu.designation if tu else '?'} ({tc.designator if tc else '?'})")
            ws6.cell(row=row6, column=1).font = Font(name="Arial", italic=True, size=9, color="6B7280")
            ws6.cell(row=row6, column=5,
                     value=f"Cable: {harness.cable_type or '—'} | Length: {harness.overall_length_m or '—'}m | Shield: {ev(harness.shield_type) or '—'}")
            ws6.cell(row=row6, column=5).font = Font(name="Arial", italic=True, size=8, color="6B7280")
            row6 += 1

            wires = db.query(Wire).filter(Wire.harness_id == harness.id).order_by(Wire.wire_number).all()
            if wires:
                wire_headers = ["Wire #", "Signal Name", "Wire Type", "Gauge",
                                "From Pin", "From Signal", "To Pin", "To Signal", "Color", "Length (m)"]
                wire_widths = [10, 20, 16, 8, 8, 16, 8, 16, 10, 10]
                _write_header(ws6, row6, wire_headers, wire_widths)
                row6 += 1

                for w in wires:
                    fp = db.query(Pin).filter(Pin.id == w.from_pin_id).first()
                    tp = db.query(Pin).filter(Pin.id == w.to_pin_id).first()
                    wtype = ev(w.wire_type)
                    fills = {}
                    if "power" in wtype:
                        fills = {3: red_fill}
                    elif "shield" in wtype or "ground" in wtype:
                        fills = {3: yellow_fill}

                    _write_row(ws6, row6, [
                        w.wire_number, w.signal_name,
                        wtype.replace("_", " "), ev(w.wire_gauge) if w.wire_gauge else "",
                        fp.pin_number if fp else "?", fp.signal_name if fp else "?",
                        tp.pin_number if tp else "?", tp.signal_name if tp else "?",
                        w.wire_color_primary or "", w.length_m,
                    ], fills=fills)
                    row6 += 1
            else:
                ws6.cell(row=row6, column=1, value="(No wires defined)").font = Font(name="Arial", italic=True, size=9, color="9CA3AF")
                row6 += 1

            row6 += 2

        # ══════════════════════════════════════
        #  Sheet 7: Signal Dictionary
        # ══════════════════════════════════════

        ws7 = wb.create_sheet("Signal Dictionary")
        ws7["A1"] = "Signal Dictionary"
        ws7["A1"].font = title_font

        sig_headers = ["Signal Name", "Type", "Direction", "Unit", "Connector", "Pin",
                       "Bus", "Harness", "Wire #", "Voltage", "Current (A)"]
        sig_widths = [20, 18, 12, 14, 10, 6, 16, 14, 10, 10, 10]
        _write_header(ws7, 3, sig_headers, sig_widths)

        # Collect all pins with their context
        all_pins = db.query(Pin).join(Connector).filter(
            Connector.project_id == project_id
        ).order_by(Pin.signal_name).all()

        row7 = 4
        for pin in all_pins:
            conn = conn_map.get(pin.connector_id)
            unit = unit_map.get(conn.unit_id) if conn else None

            # Find bus assignment
            ba = db.query(PinBusAssignment).filter(PinBusAssignment.pin_id == pin.id).first()
            bus_name = ""
            if ba:
                bus = db.query(BusDefinition).filter(BusDefinition.id == ba.bus_def_id).first()
                bus_name = bus.name if bus else ""

            # Find wire
            wire = db.query(Wire).filter(
                (Wire.from_pin_id == pin.id) | (Wire.to_pin_id == pin.id)
            ).first()
            harness_name = ""
            wire_num = ""
            if wire:
                h = next((h for h in harnesses if h.id == wire.harness_id), None)
                harness_name = h.harness_id or h.name if h else ""
                wire_num = wire.wire_number

            _write_row(ws7, row7, [
                pin.signal_name, ev(pin.signal_type).replace("_", " "),
                ev(pin.direction).replace("_", " "),
                unit.designation if unit else "?",
                conn.designator if conn else "?", pin.pin_number,
                bus_name, harness_name, wire_num,
                pin.voltage_nominal or "", pin.current_max_amps or "",
            ])
            row7 += 1

        # ══════════════════════════════════════
        #  Sheet 8: Environmental Summary
        # ══════════════════════════════════════

        ws8 = wb.create_sheet("Environmental Summary")
        ws8["A1"] = "Environmental & EMI Summary"
        ws8["A1"].font = title_font

        env_headers = ["Unit", "Category", "Standard", "Method", "Value",
                       "Min", "Max", "Status", "Auto-Generated"]
        env_widths = [14, 20, 16, 16, 10, 10, 10, 12, 12]
        _write_header(ws8, 3, env_headers, env_widths)

        row8 = 4
        for es in env_specs:
            unit = unit_map.get(es.unit_id)
            status = es.compliance_status or "untested"
            fills = {}
            if status == "pass":
                fills = {8: green_fill}
            elif status == "fail":
                fills = {8: red_fill}

            _write_row(ws8, row8, [
                unit.designation if unit else "?",
                (es.category or "").replace("_", " "),
                es.standard or "", es.test_method or "",
                es.limit_value, es.limit_min, es.limit_max,
                status, "Yes" if es.auto_generated else "No",
            ], fills=fills)
            row8 += 1

        # Also add unit-level thermal/mech summary if no env specs exist
        if not env_specs:
            row8 += 1
            ws8.cell(row=row8, column=1, value="Unit Specification Summary (from unit fields):").font = subtitle_font
            row8 += 1
            spec_headers = ["Unit", "Temp Op (°C)", "Temp Sto (°C)", "Vib (Grms)",
                            "Shock (g)", "CE102 (dBμA)", "RS103 (V/m)", "ESD (V)"]
            _write_header(ws8, row8, spec_headers, [14, 16, 16, 10, 10, 12, 12, 10])
            row8 += 1
            for u in units:
                temp_op = f"{u.temp_operating_min_c} to {u.temp_operating_max_c}" if u.temp_operating_min_c is not None else ""
                temp_sto = f"{u.temp_storage_min_c} to {u.temp_storage_max_c}" if u.temp_storage_min_c is not None else ""
                _write_row(ws8, row8, [
                    u.designation, temp_op, temp_sto,
                    u.vibration_random_grms, u.shock_mechanical_g,
                    u.emi_ce102_limit_dbua, u.emi_rs103_limit_vm, u.esd_hbm_v,
                ])
                row8 += 1

        # ══════════════════════════════════════
        #  Sheet 9: Requirements Trace
        # ══════════════════════════════════════

        ws9 = wb.create_sheet("Requirements Trace")
        ws9["A1"] = "Interface → Requirement Traceability"
        ws9["A1"].font = title_font

        req_headers = ["Entity Type", "Entity ID", "Req ID", "Req Title",
                       "Link Type", "Auto-Generated", "Template", "Status",
                       "Confidence"]
        req_widths = [14, 10, 10, 30, 12, 12, 16, 14, 10]
        _write_header(ws9, 3, req_headers, req_widths)

        # Load requirement details for display
        from app.models import Requirement
        req_map = {}
        req_ids = list({lk.requirement_id for lk in links if lk.requirement_id})
        if req_ids:
            reqs = db.query(Requirement).filter(Requirement.id.in_(req_ids)).all()
            req_map = {r.id: r for r in reqs}

        row9 = 4
        for lk in sorted(links, key=lambda x: (ev(x.entity_type), x.entity_id)):
            req = req_map.get(lk.requirement_id)
            status = ev(lk.status)
            fills = {}
            if status == "approved":
                fills = {8: green_fill}
            elif status == "rejected":
                fills = {8: red_fill}
            elif status == "pending_review":
                fills = {8: yellow_fill}

            _write_row(ws9, row9, [
                ev(lk.entity_type).replace("_", " "),
                lk.entity_id,
                req.req_id if req else "?",
                req.title if req else "?",
                ev(lk.link_type).replace("_", " "),
                "Yes" if lk.auto_generated else "No",
                lk.auto_req_template or "manual",
                status.replace("_", " "),
                lk.confidence_score,
            ], fills=fills)
            row9 += 1

        # ── Save ──
        buf = io.BytesIO()
        wb.save(buf)

        metadata = {
            "project": project.code,
            "generated_at": datetime.utcnow().isoformat(),
            "systems": len(systems),
            "units": len(units),
            "connectors": len(connectors),
            "buses": len(buses),
            "harnesses": len(harnesses),
            "interfaces": len(interfaces),
            "env_specs": len(env_specs),
            "req_links": len(links),
            "sheets": 9,
        }

        return ReportOutput(
            content=buf.getvalue(),
            filename=f"ICD_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata=metadata,
        )

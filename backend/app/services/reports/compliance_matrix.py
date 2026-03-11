"""
ASTRA — Compliance Matrix Report
==================================
File: backend/app/services/reports/compliance_matrix.py   ← NEW

Maps requirements to compliance framework controls and reports:
  - Control → Requirement mapping
  - Coverage percentage per control family
  - Gap analysis (unmapped controls)
  - Configurable: user selects which framework to map against

Mapping heuristic: scans requirement text for framework-specific
keywords.  Real deployments would store explicit control mappings
in a dedicated table; this provides a functional baseline.

Frameworks: NIST 800-53, MIL-STD-882E, DO-178C, ISO 29148
"""

import io
import re
from datetime import datetime
from collections import defaultdict

from sqlalchemy.orm import Session

from app.models import Requirement
from app.services.reports.base import ReportGenerator, ReportOutput


# ══════════════════════════════════════
#  Framework Definitions
# ══════════════════════════════════════

FRAMEWORKS: dict[str, dict[str, dict]] = {
    "nist-800-53": {
        "name": "NIST SP 800-53 Rev 5",
        "controls": {
            "AC-2": {"title": "Account Management", "keywords": ["account", "user management", "role", "access control", "rbac"]},
            "AC-7": {"title": "Unsuccessful Logon Attempts", "keywords": ["lockout", "failed login", "brute force", "login attempt"]},
            "AU-2": {"title": "Event Logging", "keywords": ["audit", "log", "event", "trail", "record"]},
            "AU-3": {"title": "Content of Audit Records", "keywords": ["audit", "timestamp", "user", "action", "detail"]},
            "AU-9": {"title": "Protection of Audit Info", "keywords": ["tamper", "immutable", "append-only", "hash chain"]},
            "IA-2": {"title": "Identification & Authentication", "keywords": ["authenticate", "login", "password", "credential", "jwt", "token", "mfa", "piv", "cac"]},
            "IA-5": {"title": "Authenticator Management", "keywords": ["password", "hash", "bcrypt", "secret", "key management"]},
            "SC-5": {"title": "Denial of Service Protection", "keywords": ["rate limit", "throttle", "dos", "denial of service"]},
            "SC-8": {"title": "Transmission Confidentiality", "keywords": ["tls", "ssl", "https", "encrypt", "transport"]},
            "SC-12": {"title": "Cryptographic Key Management", "keywords": ["key", "encryption key", "derive", "pbkdf"]},
            "SC-13": {"title": "Cryptographic Protection", "keywords": ["encrypt", "aes", "sha", "hash", "fernet", "cipher"]},
            "SC-28": {"title": "Protection of Info at Rest", "keywords": ["encrypt at rest", "field encryption", "encrypted column"]},
            "SI-11": {"title": "Error Handling", "keywords": ["error", "exception", "security header", "stack trace"]},
            "CM-3": {"title": "Configuration Change Control", "keywords": ["change control", "baseline", "version", "approval", "workflow"]},
            "CM-6": {"title": "Configuration Settings", "keywords": ["configuration", "settings", "environment", "deploy"]},
        },
    },
    "mil-std-882e": {
        "name": "MIL-STD-882E System Safety",
        "controls": {
            "4.1": {"title": "System Safety Program Plan", "keywords": ["safety program", "safety plan", "hazard"]},
            "4.2": {"title": "Hazard Analysis", "keywords": ["hazard", "risk", "severity", "probability"]},
            "4.3": {"title": "Safety Requirements", "keywords": ["safety", "shall not cause", "harm", "hazardous"]},
            "4.4": {"title": "Safety Verification", "keywords": ["safety verification", "test", "inspection", "demonstration"]},
            "4.5": {"title": "Safety Assessment", "keywords": ["safety assessment", "residual risk", "risk acceptance"]},
            "4.6": {"title": "Safety Tracking", "keywords": ["track", "monitor", "closure", "corrective action"]},
        },
    },
    "do-178c": {
        "name": "DO-178C Airborne Software",
        "controls": {
            "MB.6.1": {"title": "Requirements Standards", "keywords": ["requirement", "shall", "traceability", "verifiable"]},
            "MB.6.2": {"title": "Requirements Traceability", "keywords": ["trace", "traceability", "source", "derived"]},
            "MB.6.3": {"title": "Requirements Reviews", "keywords": ["review", "peer review", "inspection", "approval"]},
            "MB.6.4": {"title": "Requirements Verification", "keywords": ["verification", "test", "analysis", "demonstration"]},
            "MB.6.5": {"title": "Compliance & Conformance", "keywords": ["compliance", "conform", "standard", "guideline"]},
        },
    },
    "iso-29148": {
        "name": "ISO/IEC/IEEE 29148 Requirements Engineering",
        "controls": {
            "5.2.1": {"title": "Stakeholder Requirements", "keywords": ["stakeholder", "need", "concern", "expectation"]},
            "5.2.6": {"title": "Requirements Attributes", "keywords": ["priority", "status", "type", "level", "id"]},
            "6.2": {"title": "Requirement Specification", "keywords": ["specification", "statement", "shall", "rationale"]},
            "6.3": {"title": "Verification & Validation", "keywords": ["verification", "validation", "test", "acceptance"]},
            "6.4": {"title": "Requirements Management", "keywords": ["change", "baseline", "version", "history", "impact"]},
        },
    },
}


class ComplianceMatrixReport(ReportGenerator):
    name = "compliance"
    supported_formats = ["xlsx", "pdf"]

    def generate(self, project_id: int, db: Session, options: dict | None = None) -> ReportOutput:
        options = options or {}
        fmt = options.get("format", "xlsx")
        fw_key = options.get("framework", "nist-800-53")

        project = self._get_project(db, project_id)
        reqs = self._get_requirements(db, project_id)

        if fw_key not in FRAMEWORKS:
            raise ValueError(f"Unknown framework '{fw_key}'. Available: {list(FRAMEWORKS.keys())}")
        fw = FRAMEWORKS[fw_key]

        # Map controls → matching requirements
        control_map: dict[str, list[str]] = {}
        for ctrl_id, ctrl in fw["controls"].items():
            matches = []
            for req in reqs:
                text = f"{req.title} {req.statement} {req.rationale or ''}".lower()
                if any(kw in text for kw in ctrl["keywords"]):
                    matches.append(req.req_id)
            control_map[ctrl_id] = matches

        mapped = sum(1 for v in control_map.values() if v)
        total_controls = len(fw["controls"])
        coverage = round(mapped / total_controls * 100, 1) if total_controls else 0
        gaps = [cid for cid, reqs_list in control_map.items() if not reqs_list]

        summary = {
            "framework": fw["name"],
            "framework_key": fw_key,
            "total_controls": total_controls,
            "mapped_controls": mapped,
            "coverage_pct": coverage,
            "gaps": len(gaps),
            "total_requirements": len(reqs),
            "generated_at": datetime.utcnow().isoformat(),
            "project": project.code,
        }

        rows = []
        for ctrl_id in sorted(fw["controls"].keys()):
            ctrl = fw["controls"][ctrl_id]
            mapped_reqs = control_map.get(ctrl_id, [])
            rows.append({
                "control_id": ctrl_id,
                "title": ctrl["title"],
                "mapped_reqs": ", ".join(mapped_reqs) if mapped_reqs else "— GAP —",
                "count": len(mapped_reqs),
                "status": "Mapped" if mapped_reqs else "GAP",
            })

        if fmt == "xlsx":
            return self._to_xlsx(rows, summary, gaps, fw, project)
        return self._to_pdf(rows, summary, gaps, fw, project)

    def _to_xlsx(self, rows, summary, gaps, fw, project) -> ReportOutput:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Matrix"
        hf = Font(bold=True, color="FFFFFF", size=11)
        hfill = PatternFill("solid", fgColor="1F2937")
        gap_fill = PatternFill("solid", fgColor="FEE2E2")
        ok_fill = PatternFill("solid", fgColor="D1FAE5")
        thin = Border(left=Side("thin"), right=Side("thin"),
                      top=Side("thin"), bottom=Side("thin"))

        ws.merge_cells("A1:E1")
        ws["A1"] = f"Compliance Matrix — {fw['name']} — {project.code}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Coverage: {summary['coverage_pct']}% ({summary['mapped_controls']}/{summary['total_controls']} controls)"
        ws["A3"] = f"Generated: {summary['generated_at']}"

        headers = ["Control ID", "Control Title", "Mapped Requirements", "Count", "Status"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=5, column=col, value=h)
            c.font = hf
            c.fill = hfill
            c.border = thin

        for i, r in enumerate(rows, 6):
            ws.cell(row=i, column=1, value=r["control_id"]).border = thin
            ws.cell(row=i, column=2, value=r["title"]).border = thin
            ws.cell(row=i, column=3, value=r["mapped_reqs"]).border = thin
            ws.cell(row=i, column=4, value=r["count"]).border = thin
            sc = ws.cell(row=i, column=5, value=r["status"])
            sc.border = thin
            sc.fill = ok_fill if r["status"] == "Mapped" else gap_fill
            sc.font = Font(bold=True)

        for col, w in zip("ABCDE", [14, 35, 50, 8, 10]):
            ws.column_dimensions[col].width = w

        # Gap list sheet
        if gaps:
            ws2 = wb.create_sheet("Gaps")
            ws2["A1"] = "Unmapped Controls (Action Required)"
            ws2["A1"].font = Font(bold=True, size=12, color="CC0000")
            for i, g in enumerate(gaps):
                ctrl = fw["controls"][g]
                ws2.cell(row=2 + i, column=1, value=g)
                ws2.cell(row=2 + i, column=2, value=ctrl["title"])
            ws2.column_dimensions["A"].width = 14
            ws2.column_dimensions["B"].width = 40

        buf = io.BytesIO()
        wb.save(buf)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"Compliance_{summary['framework_key']}_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            metadata=summary,
        )

    def _to_pdf(self, rows, summary, gaps, fw, project) -> ReportOutput:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=LETTER)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(
            f"<b>Compliance Matrix — {fw['name']}</b>", styles["Title"]))
        elements.append(Paragraph(
            f"Project: {project.code} | Coverage: {summary['coverage_pct']}% | "
            f"Gaps: {summary['gaps']}", styles["Normal"]))
        elements.append(Spacer(1, 16))

        header = ["Control", "Title", "Requirements", "Status"]
        data = [header]
        row_bg = []
        for r in rows:
            data.append([
                r["control_id"], r["title"],
                r["mapped_reqs"][:60], r["status"],
            ])
            row_bg.append(
                colors.Color(0.82, 0.98, 0.90) if r["status"] == "Mapped"
                else colors.Color(1, 0.89, 0.89)
            )

        t = Table(data, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.12, 0.16, 0.22)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        for i, bg in enumerate(row_bg):
            style_cmds.append(("BACKGROUND", (3, i+1), (3, i+1), bg))
        t.setStyle(TableStyle(style_cmds))
        elements.append(t)

        doc.build(elements)
        return ReportOutput(
            content=buf.getvalue(),
            filename=f"Compliance_{summary['framework_key']}_{project.code}_{datetime.utcnow().strftime('%Y%m%d')}.pdf",
            content_type="application/pdf",
            metadata=summary,
        )

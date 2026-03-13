"""
ASTRA — Interface Module Import / Export Router
===================================================
File: backend/app/routers/interface_import.py
Path: C:\\Users\\Mason\\Documents\\ASTRA\\backend\\app\\routers\\interface_import.py

Endpoints:
  POST /interfaces/io/import/template    — download styled .xlsx template
  POST /interfaces/io/import/preview     — validate uploaded file, return preview
  POST /interfaces/io/import/confirm     — create all entities from validated file
  GET  /interfaces/io/export/units       — full project units/connectors/buses/messages
  GET  /interfaces/io/export/harness/{id} — single harness wire-list export
  GET  /interfaces/io/export/all-wiring  — master wiring across all harnesses
  GET  /interfaces/io/export/icd-data    — complete ICD data package (9 sheets)
"""

import io
import re
import logging
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.models import User, Project, Requirement
from app.models.interface import (
    System, Unit, Connector, Pin, BusDefinition,
    PinBusAssignment, MessageDefinition, MessageField,
    WireHarness, Wire, Interface,
    UnitEnvironmentalSpec, InterfaceRequirementLink,
)
from app.services.auth import get_current_user

try:
    from app.services.rbac import require_permission
except ImportError:
    def require_permission(action):
        return get_current_user

try:
    from app.services.audit_service import record_event as _audit
except ImportError:
    def _audit(*a, **kw):
        pass

logger = logging.getLogger("astra.interface.io")

router = APIRouter(prefix="/interfaces/io", tags=["Interface Import/Export"])


# ══════════════════════════════════════════════════════════════
#  Constants
# ══════════════════════════════════════════════════════════════

VALID_UNIT_TYPES = [
    "lru", "wru", "sru", "cca", "pcb", "backplane", "chassis", "sensor",
    "actuator", "motor", "processor", "fpga", "asic", "power_supply",
    "power_converter", "battery", "solar_panel", "transmitter", "receiver",
    "transceiver", "antenna", "waveguide", "filter_rf", "amplifier",
    "custom",
]
VALID_UNIT_STATUSES = [
    "concept", "preliminary_design", "detailed_design", "prototype",
    "engineering_model", "qualification_unit", "flight_unit", "flight_spare",
    "production", "installed", "qualified", "accepted", "operational",
]
VALID_CONNECTOR_TYPES = [
    "mil_dtl_38999_series_iii", "mil_dtl_38999_series_i", "mil_dtl_26482_series_i",
    "d_sub_9", "d_sub_15", "d_sub_25", "d_sub_37",
    "micro_d_9", "micro_d_15", "micro_d_25",
    "rj45", "usb_c", "sma", "bnc", "fiber_lc", "fiber_sc",
    "m12_4pin", "m12_8pin", "backplane_vpx", "custom",
]
VALID_GENDERS = ["male_pin", "female_socket", "hermaphroditic", "genderless"]
VALID_SIGNAL_TYPES = [
    "power_primary", "power_secondary", "power_return",
    "chassis_ground", "signal_ground",
    "signal_digital_single", "signal_digital_differential",
    "signal_analog_single", "signal_analog_differential",
    "clock_single", "clock_differential",
    "rf_signal", "discrete_input", "discrete_output",
    "serial_data", "parallel_data", "spare", "no_connect",
    "shield_overall", "shield_drain", "test_point", "custom",
]
VALID_PIN_DIRECTIONS = [
    "input", "output", "bidirectional", "tri_state",
    "power_source", "power_sink", "power_return",
    "ground", "chassis_ground", "no_connect", "spare", "custom",
]
VALID_PROTOCOLS = [
    "mil_std_1553b", "mil_std_1553a", "spacewire", "rs422", "rs485",
    "rs232", "can_2b", "canfd", "spi_mode0", "i2c_standard",
    "ethernet_100base_tx", "ethernet_1000base_t", "arinc_429",
    "usb_2_0", "jtag", "analog_4_20ma", "discrete_28v", "custom",
]
VALID_BUS_ROLES = [
    "bus_controller", "remote_terminal", "bus_monitor",
    "master", "slave", "publisher", "subscriber", "custom",
]
VALID_MSG_DIRECTIONS = [
    "transmit", "receive", "transmit_receive", "broadcast",
    "request", "response", "status",
]
VALID_DATA_TYPES = [
    "boolean", "uint8", "int8", "uint16", "int16", "uint32", "int32",
    "float32", "float64", "enum_coded", "bitfield", "bitmask",
    "char_ascii", "string_fixed", "timestamp_utc", "raw_bytes",
    "reserved", "spare", "custom",
]

# Color constants for styling
BLUE_FILL = "CCE5FF"   # Required column header
GRAY_FILL = "E8E8E8"   # Optional column header
HEADER_FONT_SIZE = 10
DATA_FONT_SIZE = 10


# ══════════════════════════════════════════════════════════════
#  Preview / Import Schemas
# ══════════════════════════════════════════════════════════════

class RowPreview(BaseModel):
    row: int
    valid: bool = True
    errors: List[str] = []
    warnings: List[str] = []
    data: dict = {}


class ImportPreviewResponse(BaseModel):
    file_name: str
    sheets_found: List[str] = []
    units: List[RowPreview] = []
    connectors: List[RowPreview] = []
    pins: List[RowPreview] = []
    buses: List[RowPreview] = []
    messages: List[RowPreview] = []
    fields: List[RowPreview] = []
    summary: dict = {}


class ImportConfirmResponse(BaseModel):
    systems_created: int = 0
    units_created: int = 0
    connectors_created: int = 0
    pins_created: int = 0
    buses_created: int = 0
    pin_assignments_created: int = 0
    messages_created: int = 0
    fields_created: int = 0
    env_specs_created: int = 0
    errors: List[str] = []
    created_ids: dict = {}


# ══════════════════════════════════════════════════════════════
#  Helpers
# ══════════════════════════════════════════════════════════════

def _ev(v) -> str:
    return v.value if hasattr(v, "value") else str(v) if v else ""


def _fuzzy_match(val: str, valid: list) -> tuple[str | None, str]:
    """Case-insensitive enum match with suggestion on failure."""
    if not val:
        return None, ""
    v = val.strip().lower().replace(" ", "_").replace("-", "_")
    if v in valid:
        return v, ""
    # Partial match
    for vv in valid:
        if v in vv or vv in v:
            return None, f"Did you mean '{vv}'?"
    return None, f"Invalid value. Options include: {', '.join(valid[:8])}..."


def _safe_float(val) -> float | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> int | None:
    if val is None or str(val).strip() == "":
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _safe_str(val, max_len: int = 255) -> str:
    if val is None:
        return ""
    return str(val).strip()[:max_len]


def _cell_val(ws, row, col) -> str:
    """Read a cell value as stripped string."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    return str(v).strip()


def _next_id(db: Session, model, project_id: int, prefix: str, id_field: str) -> str:
    max_row = db.query(model).filter(model.project_id == project_id).order_by(model.id.desc()).first()
    if max_row:
        existing_id = getattr(max_row, id_field, "") or ""
        match = re.search(r"(\d+)$", existing_id)
        next_num = (int(match.group(1)) + 1) if match else 1
    else:
        next_num = 1
    return f"{prefix}-{next_num:03d}"


# ══════════════════════════════════════════════════════════════
#  TEMPLATE GENERATION
# ══════════════════════════════════════════════════════════════

@router.post("/import/template")
def generate_import_template(
    current_user: User = Depends(get_current_user),
):
    """Generate a styled .xlsx import template with example data and validation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    blue = PatternFill(start_color=BLUE_FILL, end_color=BLUE_FILL, fill_type="solid")
    gray = PatternFill(start_color=GRAY_FILL, end_color=GRAY_FILL, fill_type="solid")
    hdr_font = Font(name="Arial", size=HEADER_FONT_SIZE, bold=True)
    data_font = Font(name="Arial", size=DATA_FONT_SIZE)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _style_sheet(ws, headers: list, required_cols: set, example_rows: list):
        """Apply headers, fills, example data, freeze, auto-filter."""
        for col_idx, (name, _width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = hdr_font
            cell.fill = blue if col_idx - 1 in required_cols else gray
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = _width

        for row_idx, row_data in enumerate(example_rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Sheet 1: Units ──
    ws_units = wb.active
    ws_units.title = "Units"
    unit_headers = [
        ("designation*", 14), ("name*", 25), ("system*", 20),
        ("part_number*", 18), ("manufacturer*", 20), ("cage_code", 10),
        ("nsn", 14), ("unit_type*", 16), ("status", 14), ("heritage", 18),
        ("mass_kg", 10), ("power_watts_nominal", 16), ("power_watts_peak", 14),
        ("voltage_input", 12), ("voltage_input_min", 14), ("voltage_input_max", 14),
        ("temp_operating_min_c", 16), ("temp_operating_max_c", 16),
        ("temp_storage_min_c", 14), ("temp_storage_max_c", 14),
        ("vibration_random_grms", 16), ("vibration_sine_g_peak", 16),
        ("shock_mechanical_g", 14), ("shock_pyrotechnic_g", 14),
        ("acceleration_max_g", 14),
        ("humidity_min_pct", 12), ("humidity_max_pct", 12),
        ("altitude_operating_max_m", 18),
        ("emi_ce102_limit_dbua", 16), ("emi_re102_limit_dbm", 16),
        ("emi_cs114_limit_dba", 16), ("emi_rs103_limit_vm", 14),
        ("esd_hbm_v", 10), ("radiation_tid_krad", 14),
        ("mtbf_hours", 12), ("design_life_years", 14),
        ("datasheet_url", 30), ("notes", 30),
    ]
    unit_required = {0, 1, 2, 3, 4, 7}  # designation, name, system, part_number, manufacturer, unit_type
    unit_examples = [
        ["RSP-100", "Radar Signal Processor", "Radar Subsystem", "RSP-100-A1",
         "Raytheon", "1RAY1", "5985-01-234-5678", "processor", "detailed_design",
         "Heritage from AN/APG-79", 12.5, 85.0, 120.0, "28VDC", 22.0, 32.0,
         -40, 71, -62, 85, 14.1, 10.0, 40, 300, 15.0,
         10, 95, 15240, 72, 48, 31, 10, 2000, 50,
         25000, 15, "https://example.com/rsp100", "Flight-qualified unit"],
        ["ANT-200", "Phased Array Antenna", "Radar Subsystem", "ANT-200-B2",
         "Northrop Grumman", "2NGR2", "", "antenna", "preliminary_design",
         "", 45.0, 250.0, 400.0, "28VDC", 22.0, 32.0,
         -54, 80, -62, 85, 20.0, 15.0, 75, 500, 20.0,
         5, 100, 21000, 80, 55, 37, 15, 4000, 100,
         15000, 20, "", "AESA antenna assembly"],
    ]
    _style_sheet(ws_units, unit_headers, unit_required, unit_examples)

    # Data validation for unit_type and status
    dv_type = DataValidation(type="list", formula1=f'"{",".join(VALID_UNIT_TYPES[:25])}"', allow_blank=False)
    dv_type.error = "Invalid unit type"
    dv_type.prompt = "Select unit type"
    ws_units.add_data_validation(dv_type)
    dv_type.add(f"H2:H1000")

    dv_status = DataValidation(type="list", formula1=f'"{",".join(VALID_UNIT_STATUSES)}"', allow_blank=True)
    ws_units.add_data_validation(dv_status)
    dv_status.add(f"I2:I1000")

    # ── Sheet 2: Connectors (one row per PIN) ──
    ws_conn = wb.create_sheet("Connectors")
    conn_headers = [
        ("unit_designation*", 16), ("connector_designator*", 18), ("connector_name", 20),
        ("connector_type*", 22), ("gender*", 14), ("shell_size", 10),
        ("insert_arrangement", 14), ("total_contacts*", 14), ("mil_spec", 20),
        ("keying", 10), ("mounting", 14),
        ("pin_number*", 10), ("signal_name*", 22), ("signal_type*", 20),
        ("direction*", 14), ("voltage_nominal", 12), ("current_max_amps", 14),
        ("impedance_ohms", 12), ("description", 30),
    ]
    conn_required = {0, 1, 3, 4, 7, 11, 12, 13, 14}
    conn_examples = [
        ["RSP-100", "J1", "1553 Bus A", "mil_dtl_38999_series_iii", "female_socket",
         "13", "13-35", 22, "M38999/26WB35SN", "N", "panel_mount",
         "A", "1553A_HI", "signal_digital_differential", "bidirectional",
         "", "", "78", "MIL-STD-1553B Bus A High"],
        ["RSP-100", "J1", "1553 Bus A", "mil_dtl_38999_series_iii", "female_socket",
         "13", "13-35", 22, "", "", "",
         "B", "1553A_LO", "signal_digital_differential", "bidirectional",
         "", "", "78", "MIL-STD-1553B Bus A Low"],
        ["RSP-100", "J1", "1553 Bus A", "mil_dtl_38999_series_iii", "female_socket",
         "13", "13-35", 22, "", "", "",
         "C", "1553A_SHIELD", "shield_drain", "ground",
         "", "", "", "Bus A shield drain"],
        ["RSP-100", "J1", "1553 Bus A", "", "", "", "", "", "", "", "",
         "D", "1553B_HI", "signal_digital_differential", "bidirectional",
         "", "", "78", "MIL-STD-1553B Bus B High"],
        ["RSP-100", "J1", "", "", "", "", "", "", "", "", "",
         "E", "1553B_LO", "signal_digital_differential", "bidirectional",
         "", "", "78", "Bus B Low"],
        ["RSP-100", "J2", "Power Input", "mil_dtl_38999_series_iii", "female_socket",
         "11", "11-2", 4, "", "A", "panel_mount",
         "1", "PWR_28V_PRI", "power_primary", "power_sink",
         "28VDC", "5.0", "", "Primary 28V input"],
        ["RSP-100", "J2", "Power Input", "", "", "", "", "", "", "", "",
         "2", "PWR_28V_RTN", "power_return", "power_return",
         "", "5.0", "", "Primary 28V return"],
        ["RSP-100", "J2", "", "", "", "", "", "", "", "", "",
         "3", "CHASSIS_GND", "chassis_ground", "chassis_ground",
         "", "", "", "Chassis ground stud"],
        ["RSP-100", "J2", "", "", "", "", "", "", "", "", "",
         "4", "SPARE_J2_4", "spare", "no_connect",
         "", "", "", "Spare contact"],
        ["RSP-100", "J3", "Ethernet", "rj45", "female_socket",
         "", "", 8, "", "", "panel_mount",
         "1", "ETH_TX_P", "signal_digital_differential", "output",
         "", "", "100", "Ethernet TX+"],
    ]
    _style_sheet(ws_conn, conn_headers, conn_required, conn_examples)

    # ── Sheet 3: Buses ──
    ws_bus = wb.create_sheet("Buses")
    bus_headers = [
        ("unit_designation*", 16), ("bus_name*", 22), ("protocol*", 20),
        ("bus_role*", 16), ("bus_address", 12), ("data_rate", 14),
        ("word_size_bits", 12), ("bus_name_network", 18), ("redundancy", 14),
        ("pin_assignments", 40),
    ]
    bus_required = {0, 1, 2, 3}
    bus_examples = [
        ["RSP-100", "1553 Bus A", "mil_std_1553b", "remote_terminal", "RT05",
         "1 Mbps", 16, "MUX_BUS_A", "dual_standby",
         "J1:A(data_positive),J1:B(data_negative),J1:C(shield)"],
        ["RSP-100", "Ethernet LAN", "ethernet_100base_tx", "peer", "",
         "100 Mbps", 8, "MGMT_LAN", "none",
         "J3:1(tx_positive),J3:2(tx_negative),J3:3(rx_positive),J3:6(rx_negative)"],
    ]
    _style_sheet(ws_bus, bus_headers, bus_required, bus_examples)

    # ── Sheet 4: Messages (one row per FIELD) ──
    ws_msg = wb.create_sheet("Messages")
    msg_headers = [
        ("unit_designation*", 16), ("bus_name*", 20), ("msg_label*", 22),
        ("msg_mnemonic", 12), ("direction*", 14), ("subaddress", 10),
        ("word_count", 10), ("message_id_hex", 12),
        ("rate_hz", 8), ("latency_max_ms", 12), ("priority", 10), ("scheduling", 20),
        ("field_name*", 22), ("field_label", 18), ("data_type*", 14),
        ("word_number", 10), ("bit_offset", 10), ("bit_length*", 10),
        ("unit_of_measure", 12), ("scale_factor", 10), ("offset_value", 10),
        ("min_value", 10), ("max_value", 10), ("default_value", 10),
        ("enum_values", 30),
    ]
    msg_required = {0, 1, 2, 4, 12, 14, 17}
    msg_examples = [
        ["RSP-100", "1553 Bus A", "Target Track", "TGT_TRK", "transmit", 5, 16, "0x0500",
         50, 20, "mission_critical", "periodic_synchronous",
         "target_id", "Target ID", "uint16", 1, 0, 16, "", 1, 0, 0, 65534, "0", ""],
        ["RSP-100", "1553 Bus A", "Target Track", "", "", "", "", "",
         "", "", "", "",
         "range_m", "Range", "float32", 2, 0, 32, "meters", 0.1, 0, 0, 300000, "0", ""],
        ["RSP-100", "1553 Bus A", "Target Track", "", "", "", "", "",
         "", "", "", "",
         "azimuth_deg", "Azimuth", "float32", 4, 0, 32, "degrees", 0.01, 0, 0, 360, "0", ""],
        ["RSP-100", "1553 Bus A", "Target Track", "", "", "", "", "",
         "", "", "", "",
         "elevation_deg", "Elevation", "float32", 6, 0, 32, "degrees", 0.01, -90, 90, "0", ""],
        ["RSP-100", "1553 Bus A", "Target Track", "", "", "", "", "",
         "", "", "", "",
         "track_quality", "Quality", "enum_coded", 8, 0, 8, "", 1, 0, 0, 4, "0",
         "0=NONE;1=TENTATIVE;2=FIRM;3=COASTING;4=LOST"],
        ["RSP-100", "1553 Bus A", "Target Track", "", "", "", "", "",
         "", "", "", "",
         "spare_w9", "Spare", "spare", 9, 0, 128, "", 1, 0, "", "", "0", ""],
    ]
    _style_sheet(ws_msg, msg_headers, msg_required, msg_examples)

    # Write to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=astra_interface_template.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  IMPORT PREVIEW
# ══════════════════════════════════════════════════════════════

@router.post("/import/preview", response_model=ImportPreviewResponse)
async def preview_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Parse and validate an uploaded .xlsx — returns preview with errors/warnings."""
    from openpyxl import load_workbook

    content = await file.read()
    if not content:
        raise HTTPException(400, "Empty file")

    filename = file.filename or "upload.xlsx"
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be .xlsx")

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    sheets_found = wb.sheetnames

    result = ImportPreviewResponse(file_name=filename, sheets_found=sheets_found)
    systems_to_create: set = set()
    unit_designations: set = set()
    connector_keys: set = set()  # (unit_des, conn_des)
    pin_keys: set = set()  # (unit_des, conn_des, pin_num)

    # ── Parse Units sheet ──
    if "Units" in wb.sheetnames:
        ws = wb["Units"]
        headers = [_safe_str(ws.cell(1, c).value).lower().replace("*", "").strip()
                   for c in range(1, ws.max_column + 1)]
        for row_idx in range(2, ws.max_row + 1):
            vals = {headers[c]: _safe_str(ws.cell(row_idx, c + 1).value)
                    for c in range(len(headers))}
            if not any(vals.values()):
                continue

            rp = RowPreview(row=row_idx, data=vals)

            # Required fields
            for req in ["designation", "name", "system", "part_number", "manufacturer", "unit_type"]:
                if not vals.get(req):
                    rp.errors.append(f"Missing required: {req}")
                    rp.valid = False

            # Validate unit_type
            if vals.get("unit_type"):
                matched, hint = _fuzzy_match(vals["unit_type"], VALID_UNIT_TYPES)
                if not matched:
                    rp.warnings.append(f"unit_type '{vals['unit_type']}' unknown. {hint}")

            # Track systems and designations
            if vals.get("system"):
                systems_to_create.add(vals["system"])
            if vals.get("designation"):
                if vals["designation"] in unit_designations:
                    rp.errors.append(f"Duplicate designation: {vals['designation']}")
                    rp.valid = False
                unit_designations.add(vals["designation"])

            result.units.append(rp)

    # ── Parse Connectors sheet ──
    if "Connectors" in wb.sheetnames:
        ws = wb["Connectors"]
        headers = [_safe_str(ws.cell(1, c).value).lower().replace("*", "").strip()
                   for c in range(1, ws.max_column + 1)]
        for row_idx in range(2, ws.max_row + 1):
            vals = {headers[c]: _safe_str(ws.cell(row_idx, c + 1).value)
                    for c in range(len(headers))}
            if not any(vals.values()):
                continue

            unit_des = vals.get("unit_designation", "")
            conn_des = vals.get("connector_designator", "")
            pin_num = vals.get("pin_number", "")

            # Connector-level validation (only for first pin row of connector)
            conn_key = (unit_des, conn_des)
            if conn_key not in connector_keys and conn_des:
                rp_conn = RowPreview(row=row_idx, data={
                    "unit_designation": unit_des,
                    "designator": conn_des,
                    "type": vals.get("connector_type", ""),
                    "gender": vals.get("gender", ""),
                })
                if not unit_des:
                    rp_conn.errors.append("Missing unit_designation")
                    rp_conn.valid = False
                if vals.get("connector_type"):
                    m, h = _fuzzy_match(vals["connector_type"], VALID_CONNECTOR_TYPES)
                    if not m:
                        rp_conn.warnings.append(f"connector_type unknown. {h}")
                result.connectors.append(rp_conn)
                connector_keys.add(conn_key)

            # Pin-level validation
            rp_pin = RowPreview(row=row_idx, data={
                "unit": unit_des, "connector": conn_des,
                "pin_number": pin_num, "signal_name": vals.get("signal_name", ""),
            })
            for req in ["pin_number", "signal_name", "signal_type", "direction"]:
                if not vals.get(req):
                    rp_pin.errors.append(f"Missing required: {req}")
                    rp_pin.valid = False

            pin_key = (unit_des, conn_des, pin_num)
            if pin_key in pin_keys and pin_num:
                rp_pin.errors.append(f"Duplicate pin {pin_num} on {conn_des}")
                rp_pin.valid = False
            pin_keys.add(pin_key)

            result.pins.append(rp_pin)

    # ── Parse Buses sheet ──
    if "Buses" in wb.sheetnames:
        ws = wb["Buses"]
        headers = [_safe_str(ws.cell(1, c).value).lower().replace("*", "").strip()
                   for c in range(1, ws.max_column + 1)]
        for row_idx in range(2, ws.max_row + 1):
            vals = {headers[c]: _safe_str(ws.cell(row_idx, c + 1).value)
                    for c in range(len(headers))}
            if not any(vals.values()):
                continue
            rp = RowPreview(row=row_idx, data=vals)
            for req in ["unit_designation", "bus_name", "protocol", "bus_role"]:
                if not vals.get(req):
                    rp.errors.append(f"Missing required: {req}")
                    rp.valid = False
            result.buses.append(rp)

    # ── Parse Messages sheet ──
    if "Messages" in wb.sheetnames:
        ws = wb["Messages"]
        headers = [_safe_str(ws.cell(1, c).value).lower().replace("*", "").strip()
                   for c in range(1, ws.max_column + 1)]
        msg_keys: set = set()
        for row_idx in range(2, ws.max_row + 1):
            vals = {headers[c]: _safe_str(ws.cell(row_idx, c + 1).value)
                    for c in range(len(headers))}
            if not any(vals.values()):
                continue

            msg_label = vals.get("msg_label", "")
            msg_key = (vals.get("unit_designation", ""), vals.get("bus_name", ""), msg_label)

            if msg_key not in msg_keys and msg_label:
                rp_msg = RowPreview(row=row_idx, data={
                    "unit": vals.get("unit_designation", ""),
                    "bus": vals.get("bus_name", ""),
                    "label": msg_label,
                    "direction": vals.get("direction", ""),
                })
                for req in ["unit_designation", "bus_name", "msg_label", "direction"]:
                    if not vals.get(req):
                        rp_msg.errors.append(f"Missing required: {req}")
                        rp_msg.valid = False
                result.messages.append(rp_msg)
                msg_keys.add(msg_key)

            # Field-level
            rp_f = RowPreview(row=row_idx, data={
                "msg": msg_label,
                "field_name": vals.get("field_name", ""),
                "data_type": vals.get("data_type", ""),
                "bit_length": vals.get("bit_length", ""),
            })
            for req in ["field_name", "data_type", "bit_length"]:
                if not vals.get(req):
                    rp_f.errors.append(f"Missing required: {req}")
                    rp_f.valid = False
            result.fields.append(rp_f)

    wb.close()

    # Build summary
    result.summary = {
        "total_units": len(result.units),
        "valid_units": sum(1 for u in result.units if u.valid),
        "total_connectors": len(result.connectors),
        "valid_connectors": sum(1 for c in result.connectors if c.valid),
        "total_pins": len(result.pins),
        "valid_pins": sum(1 for p in result.pins if p.valid),
        "error_pins": sum(1 for p in result.pins if not p.valid),
        "total_buses": len(result.buses),
        "total_messages": len(result.messages),
        "total_fields": len(result.fields),
        "systems_to_create": sorted(systems_to_create),
    }

    return result


# ══════════════════════════════════════════════════════════════
#  IMPORT CONFIRM
# ══════════════════════════════════════════════════════════════

@router.post("/import/confirm", response_model=ImportConfirmResponse)
async def confirm_import(
    project_id: int = Query(...),
    file: UploadFile = File(...),
    request=None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("interfaces.create")),
):
    """Create all entities from validated upload file."""
    from openpyxl import load_workbook

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    content = await file.read()
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)

    resp = ImportConfirmResponse(created_ids={})
    errors: list = []

    # Lookup maps built during creation
    system_map: dict[str, int] = {}  # name → id
    unit_map: dict[str, int] = {}    # designation → id
    connector_map: dict[str, int] = {}  # "unit_des|conn_des" → id
    bus_map: dict[str, int] = {}     # "unit_des|bus_name" → id
    msg_map: dict[str, int] = {}     # "unit_des|bus_name|label" → id

    # Pre-load existing systems and units
    for s in db.query(System).filter(System.project_id == project_id).all():
        system_map[s.name] = s.id
    for u in db.query(Unit).filter(Unit.project_id == project_id).all():
        unit_map[u.designation] = u.id

    def _read_sheet(name):
        if name not in wb.sheetnames:
            return [], []
        ws = wb[name]
        hdrs = [_safe_str(ws.cell(1, c).value).lower().replace("*", "").strip()
                for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = {hdrs[c]: ws.cell(r, c + 1).value for c in range(len(hdrs))}
            if any(v is not None for v in vals.values()):
                rows.append(vals)
        return hdrs, rows

    # ── 1. Systems (auto-create from unit references) ──
    _, unit_rows = _read_sheet("Units")
    for row in unit_rows:
        sys_name = _safe_str(row.get("system"))
        if sys_name and sys_name not in system_map:
            sys_id_str = _next_id(db, System, project_id, "SYS", "system_id")
            sys_obj = System(
                system_id=sys_id_str, name=sys_name, system_type="subsystem",
                project_id=project_id, owner_id=current_user.id,
            )
            db.add(sys_obj)
            db.flush()
            system_map[sys_name] = sys_obj.id
            resp.systems_created += 1

    # ── 2. Units ──
    for row in unit_rows:
        des = _safe_str(row.get("designation"))
        if not des or des in unit_map:
            continue
        sys_name = _safe_str(row.get("system"))
        sys_id = system_map.get(sys_name)
        if not sys_id:
            errors.append(f"System '{sys_name}' not found for unit '{des}'")
            continue

        uid_str = _next_id(db, Unit, project_id, "UNIT", "unit_id")
        unit = Unit(
            unit_id=uid_str,
            designation=des,
            name=_safe_str(row.get("name")),
            part_number=_safe_str(row.get("part_number")) or "TBD",
            manufacturer=_safe_str(row.get("manufacturer")) or "TBD",
            unit_type=_safe_str(row.get("unit_type")).lower().replace(" ", "_") or "custom",
            status=_safe_str(row.get("status")).lower().replace(" ", "_") or "concept",
            heritage=_safe_str(row.get("heritage")),
            cage_code=_safe_str(row.get("cage_code")),
            nsn=_safe_str(row.get("nsn")),
            system_id=sys_id,
            project_id=project_id,
            # Specs
            mass_kg=_safe_float(row.get("mass_kg")),
            power_watts_nominal=_safe_float(row.get("power_watts_nominal")),
            power_watts_peak=_safe_float(row.get("power_watts_peak")),
            voltage_input_nominal=_safe_str(row.get("voltage_input")),
            voltage_input_min=_safe_float(row.get("voltage_input_min")),
            voltage_input_max=_safe_float(row.get("voltage_input_max")),
            temp_operating_min_c=_safe_float(row.get("temp_operating_min_c")),
            temp_operating_max_c=_safe_float(row.get("temp_operating_max_c")),
            temp_storage_min_c=_safe_float(row.get("temp_storage_min_c")),
            temp_storage_max_c=_safe_float(row.get("temp_storage_max_c")),
            vibration_random_grms=_safe_float(row.get("vibration_random_grms")),
            vibration_sine_g_peak=_safe_float(row.get("vibration_sine_g_peak")),
            shock_mechanical_g=_safe_float(row.get("shock_mechanical_g")),
            shock_pyrotechnic_g=_safe_float(row.get("shock_pyrotechnic_g")),
            acceleration_max_g=_safe_float(row.get("acceleration_max_g")),
            humidity_min_pct=_safe_float(row.get("humidity_min_pct")),
            humidity_max_pct=_safe_float(row.get("humidity_max_pct")),
            altitude_operating_max_m=_safe_float(row.get("altitude_operating_max_m")),
            emi_ce102_limit_dbua=_safe_float(row.get("emi_ce102_limit_dbua")),
            emi_re102_limit_dbm=_safe_float(row.get("emi_re102_limit_dbm")),
            emi_cs114_limit_dba=_safe_float(row.get("emi_cs114_limit_dba")),
            emi_rs103_limit_vm=_safe_float(row.get("emi_rs103_limit_vm")),
            esd_hbm_v=_safe_float(row.get("esd_hbm_v")),
            radiation_tid_krad=_safe_float(row.get("radiation_tid_krad")),
            mtbf_hours=_safe_float(row.get("mtbf_hours")),
            design_life_years=_safe_float(row.get("design_life_years")),
            datasheet_url=_safe_str(row.get("datasheet_url"), 500),
            notes=_safe_str(row.get("notes"), 2000),
        )
        db.add(unit)
        db.flush()
        unit_map[des] = unit.id
        resp.units_created += 1

    # ── 3+4. Connectors + Pins ──
    _, conn_rows = _read_sheet("Connectors")
    for row in conn_rows:
        unit_des = _safe_str(row.get("unit_designation"))
        conn_des = _safe_str(row.get("connector_designator"))
        pin_num = _safe_str(row.get("pin_number"))

        unit_id = unit_map.get(unit_des)
        if not unit_id:
            errors.append(f"Unit '{unit_des}' not found for connector row")
            continue

        # Create connector if needed
        ck = f"{unit_des}|{conn_des}"
        if ck not in connector_map and conn_des:
            conn_id_str = _next_id(db, Connector, project_id, "CONN", "connector_id")
            conn = Connector(
                connector_id=conn_id_str,
                designator=conn_des,
                name=_safe_str(row.get("connector_name")),
                connector_type=_safe_str(row.get("connector_type")).lower().replace(" ", "_") or "custom",
                gender=_safe_str(row.get("gender")).lower().replace(" ", "_") or "female_socket",
                shell_size=_safe_str(row.get("shell_size")),
                insert_arrangement=_safe_str(row.get("insert_arrangement")),
                total_contacts=_safe_int(row.get("total_contacts")) or 0,
                mil_spec=_safe_str(row.get("mil_spec")),
                keying=_safe_str(row.get("keying")),
                mounting=_safe_str(row.get("mounting")).lower().replace(" ", "_") or None,
                unit_id=unit_id,
                project_id=project_id,
            )
            db.add(conn)
            db.flush()
            connector_map[ck] = conn.id
            resp.connectors_created += 1

        # Create pin
        conn_id = connector_map.get(ck)
        if conn_id and pin_num:
            pin = Pin(
                connector_id=conn_id,
                pin_number=pin_num,
                signal_name=_safe_str(row.get("signal_name")) or f"SPARE_{pin_num}",
                signal_type=_safe_str(row.get("signal_type")).lower().replace(" ", "_") or "spare",
                direction=_safe_str(row.get("direction")).lower().replace(" ", "_") or "no_connect",
                voltage_nominal=_safe_str(row.get("voltage_nominal")),
                current_max_amps=_safe_float(row.get("current_max_amps")),
                impedance_ohms=_safe_float(row.get("impedance_ohms")),
                description=_safe_str(row.get("description")),
            )
            db.add(pin)
            resp.pins_created += 1

    db.flush()

    # ── 5+6. Buses + Pin Assignments ──
    _, bus_rows = _read_sheet("Buses")
    for row in bus_rows:
        unit_des = _safe_str(row.get("unit_designation"))
        bus_name = _safe_str(row.get("bus_name"))
        unit_id = unit_map.get(unit_des)
        if not unit_id:
            errors.append(f"Unit '{unit_des}' not found for bus '{bus_name}'")
            continue

        bk = f"{unit_des}|{bus_name}"
        if bk not in bus_map:
            bd_id_str = _next_id(db, BusDefinition, project_id, "BUS", "bus_def_id")
            bd = BusDefinition(
                bus_def_id=bd_id_str,
                name=bus_name,
                protocol=_safe_str(row.get("protocol")).lower().replace(" ", "_") or "custom",
                bus_role=_safe_str(row.get("bus_role")).lower().replace(" ", "_") or "custom",
                bus_address=_safe_str(row.get("bus_address")),
                data_rate=_safe_str(row.get("data_rate")),
                word_size_bits=_safe_int(row.get("word_size_bits")),
                bus_name_network=_safe_str(row.get("bus_name_network")),
                redundancy=_safe_str(row.get("redundancy")).lower().replace(" ", "_") or "none",
                unit_id=unit_id,
                project_id=project_id,
            )
            db.add(bd)
            db.flush()
            bus_map[bk] = bd.id
            resp.buses_created += 1

        # Parse pin assignments: "J1:A(data_positive),J1:B(data_negative)"
        pa_str = _safe_str(row.get("pin_assignments"))
        if pa_str:
            for entry in pa_str.split(","):
                entry = entry.strip()
                m = re.match(r"(\w+):(\w+)\((\w+)\)", entry)
                if not m:
                    continue
                pa_conn_des, pa_pin_num, pa_role = m.group(1), m.group(2), m.group(3)
                pa_ck = f"{unit_des}|{pa_conn_des}"
                pa_conn_id = connector_map.get(pa_ck)
                if not pa_conn_id:
                    continue
                pa_pin = db.query(Pin).filter(
                    Pin.connector_id == pa_conn_id, Pin.pin_number == pa_pin_num
                ).first()
                if pa_pin:
                    pa = PinBusAssignment(
                        pin_id=pa_pin.id,
                        bus_def_id=bus_map[bk],
                        pin_role=pa_role,
                    )
                    db.add(pa)
                    resp.pin_assignments_created += 1

    db.flush()

    # ── 7+8. Messages + Fields ──
    _, msg_rows = _read_sheet("Messages")
    current_msg_key = None
    for row in msg_rows:
        unit_des = _safe_str(row.get("unit_designation"))
        bus_name = _safe_str(row.get("bus_name"))
        msg_label = _safe_str(row.get("msg_label"))

        # Inherit from previous row if blank (continuation rows)
        if not unit_des and current_msg_key:
            unit_des = current_msg_key[0]
        if not bus_name and current_msg_key:
            bus_name = current_msg_key[1]
        if not msg_label and current_msg_key:
            msg_label = current_msg_key[2]

        unit_id = unit_map.get(unit_des)
        bk = f"{unit_des}|{bus_name}"
        bus_id = bus_map.get(bk)

        if not unit_id or not bus_id:
            if msg_label:
                errors.append(f"Unit/bus not found for message '{msg_label}'")
            continue

        mk = (unit_des, bus_name, msg_label)

        # Create message if new
        mk_str = f"{unit_des}|{bus_name}|{msg_label}"
        if mk_str not in msg_map and msg_label:
            mid_str = _next_id(db, MessageDefinition, project_id, "MSG", "msg_def_id")
            msg = MessageDefinition(
                msg_def_id=mid_str,
                label=msg_label,
                mnemonic=_safe_str(row.get("msg_mnemonic")),
                direction=_safe_str(row.get("direction")).lower().replace(" ", "_") or "transmit",
                subaddress=_safe_int(row.get("subaddress")),
                word_count=_safe_int(row.get("word_count")),
                message_id_hex=_safe_str(row.get("message_id_hex")),
                rate_hz=_safe_float(row.get("rate_hz")),
                latency_max_ms=_safe_float(row.get("latency_max_ms")),
                priority=_safe_str(row.get("priority")).lower().replace(" ", "_") or "medium",
                scheduling=_safe_str(row.get("scheduling")).lower().replace(" ", "_") or None,
                bus_def_id=bus_id,
                unit_id=unit_id,
                project_id=project_id,
            )
            db.add(msg)
            db.flush()
            msg_map[mk_str] = msg.id
            resp.messages_created += 1
            current_msg_key = mk

        # Create field
        field_name = _safe_str(row.get("field_name"))
        if field_name and mk_str in msg_map:
            enum_str = _safe_str(row.get("enum_values"))
            enum_dict = None
            if enum_str:
                try:
                    enum_dict = dict(pair.split("=", 1) for pair in enum_str.split(";") if "=" in pair)
                except Exception:
                    pass

            field = MessageField(
                message_id=msg_map[mk_str],
                field_name=field_name,
                label=_safe_str(row.get("field_label")),
                data_type=_safe_str(row.get("data_type")).lower().replace(" ", "_") or "uint16",
                word_number=_safe_int(row.get("word_number")),
                bit_offset=_safe_int(row.get("bit_offset")),
                bit_length=_safe_int(row.get("bit_length")) or 16,
                unit_of_measure=_safe_str(row.get("unit_of_measure")),
                scale_factor=_safe_float(row.get("scale_factor")),
                offset_value=_safe_float(row.get("offset_value")),
                min_value=_safe_float(row.get("min_value")),
                max_value=_safe_float(row.get("max_value")),
                default_value=_safe_str(row.get("default_value")),
                enum_values=enum_dict,
                field_order=resp.fields_created + 1,
            )
            db.add(field)
            resp.fields_created += 1

    wb.close()

    resp.errors = errors
    db.commit()

    _audit(db, "interface.imported", "project", project_id, current_user.id,
           {"units": resp.units_created, "connectors": resp.connectors_created,
            "pins": resp.pins_created, "buses": resp.buses_created,
            "messages": resp.messages_created, "fields": resp.fields_created},
           project_id=project_id, request=request)

    return resp


# ══════════════════════════════════════════════════════════════
#  EXPORT — Units (full project)
# ══════════════════════════════════════════════════════════════

@router.get("/export/units")
def export_units(
    project_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all units, connectors, buses, messages to styled .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    wb = Workbook()
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_headers(ws, headers):
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        ws.freeze_panes = "A2"

    def _write_row(ws, row_idx, values):
        for i, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=i, value=v)
            cell.font = data_font
            cell.border = thin_border

    # ── Sheet 1: Units ──
    ws = wb.active
    ws.title = "Units"
    unit_hdrs = [
        "Unit ID", "Designation", "Name", "System", "Part Number", "Manufacturer",
        "Type", "Status", "Mass (kg)", "Power Nom (W)", "Power Peak (W)",
        "Temp Op Min (°C)", "Temp Op Max (°C)", "Vib Random (gRMS)",
        "Shock Mech (g)", "MTBF (hrs)", "Design Life (yr)",
    ]
    _write_headers(ws, unit_hdrs)
    units = db.query(Unit).filter(Unit.project_id == project_id).order_by(Unit.designation).all()
    for idx, u in enumerate(units, 2):
        sys_obj = db.query(System).filter(System.id == u.system_id).first()
        _write_row(ws, idx, [
            u.unit_id, u.designation, u.name, sys_obj.name if sys_obj else "",
            u.part_number, u.manufacturer, _ev(u.unit_type), _ev(u.status),
            u.mass_kg, u.power_watts_nominal, u.power_watts_peak,
            u.temp_operating_min_c, u.temp_operating_max_c,
            u.vibration_random_grms, u.shock_mechanical_g,
            u.mtbf_hours, u.design_life_years,
        ])

    # ── Sheet 2: Connectors + Pins ──
    ws2 = wb.create_sheet("Connectors")
    conn_hdrs = [
        "Unit", "Connector", "Type", "Gender", "Shell Size", "Total Contacts",
        "Pin #", "Signal Name", "Signal Type", "Direction",
        "Voltage", "Current (A)", "Impedance (Ω)", "Description",
    ]
    _write_headers(ws2, conn_hdrs)
    row_idx = 2
    for u in units:
        conns = db.query(Connector).filter(Connector.unit_id == u.id).order_by(Connector.designator).all()
        for c in conns:
            pins = db.query(Pin).filter(Pin.connector_id == c.id).order_by(Pin.pin_number).all()
            for p in pins:
                _write_row(ws2, row_idx, [
                    u.designation, c.designator, _ev(c.connector_type), _ev(c.gender),
                    c.shell_size, c.total_contacts,
                    p.pin_number, p.signal_name, _ev(p.signal_type), _ev(p.direction),
                    p.voltage_nominal, p.current_max_amps, p.impedance_ohms, p.description,
                ])
                row_idx += 1

    # ── Sheet 3: Buses ──
    ws3 = wb.create_sheet("Buses")
    bus_hdrs = [
        "Unit", "Bus Name", "Protocol", "Role", "Address", "Data Rate",
        "Word Size", "Network", "Redundancy", "Messages",
    ]
    _write_headers(ws3, bus_hdrs)
    row_idx = 2
    for u in units:
        buses = db.query(BusDefinition).filter(BusDefinition.unit_id == u.id).all()
        for bd in buses:
            mc = db.query(func.count(MessageDefinition.id)).filter(
                MessageDefinition.bus_def_id == bd.id
            ).scalar()
            _write_row(ws3, row_idx, [
                u.designation, bd.name, _ev(bd.protocol), _ev(bd.bus_role),
                bd.bus_address, bd.data_rate, bd.word_size_bits,
                bd.bus_name_network, _ev(bd.redundancy), mc,
            ])
            row_idx += 1

    # ── Sheet 4: Messages + Fields ──
    ws4 = wb.create_sheet("Messages")
    msg_hdrs = [
        "Unit", "Bus", "Message", "Mnemonic", "Direction", "SA", "Words",
        "Rate (Hz)", "Priority",
        "Field", "Data Type", "Word#", "Bit Offset", "Bit Length",
        "UoM", "Scale", "Min", "Max",
    ]
    _write_headers(ws4, msg_hdrs)
    row_idx = 2
    for u in units:
        buses = db.query(BusDefinition).filter(BusDefinition.unit_id == u.id).all()
        for bd in buses:
            msgs = db.query(MessageDefinition).filter(MessageDefinition.bus_def_id == bd.id).all()
            for msg in msgs:
                fields = db.query(MessageField).filter(
                    MessageField.message_id == msg.id
                ).order_by(MessageField.field_order).all()
                for f in fields:
                    _write_row(ws4, row_idx, [
                        u.designation, bd.name, msg.label, msg.mnemonic,
                        _ev(msg.direction), msg.subaddress, msg.word_count,
                        msg.rate_hz, _ev(msg.priority),
                        f.field_name, _ev(f.data_type), f.word_number,
                        f.bit_offset, f.bit_length,
                        f.unit_of_measure, f.scale_factor, f.min_value, f.max_value,
                    ])
                    row_idx += 1

    # Auto-fit columns
    for ws_sheet in wb.worksheets:
        for col in ws_sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws_sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=astra_{project.code}_units.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  EXPORT — Single Harness
# ══════════════════════════════════════════════════════════════

@router.get("/export/harness/{harness_id}")
def export_harness(
    harness_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export a single harness with wire list and pin-to-pin mapping."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    harness = db.query(WireHarness).filter(WireHarness.id == harness_id).first()
    if not harness:
        raise HTTPException(404, "Harness not found")

    wb = Workbook()
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _hdr(ws, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(1, i, h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border
        ws.freeze_panes = "A2"

    def _row(ws, r, vals):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = data_font
            c.border = thin_border

    fu = db.query(Unit).filter(Unit.id == harness.from_unit_id).first()
    fc = db.query(Connector).filter(Connector.id == harness.from_connector_id).first()
    tu = db.query(Unit).filter(Unit.id == harness.to_unit_id).first()
    tc = db.query(Connector).filter(Connector.id == harness.to_connector_id).first()

    # ── Summary sheet ──
    ws1 = wb.active
    ws1.title = "Harness Summary"
    summary_data = [
        ("Harness ID", harness.harness_id), ("Name", harness.name),
        ("Status", _ev(harness.status)),
        ("From Unit", fu.designation if fu else ""), ("From Connector", fc.designator if fc else ""),
        ("To Unit", tu.designation if tu else ""), ("To Connector", tc.designator if tc else ""),
        ("Cable Type", harness.cable_type), ("Length (m)", harness.overall_length_m),
        ("Shield Type", _ev(harness.shield_type)),
        ("Drawing", harness.drawing_number), ("Revision", harness.drawing_revision),
    ]
    for r, (label, val) in enumerate(summary_data, 1):
        ws1.cell(r, 1, label).font = Font(name="Arial", bold=True)
        ws1.cell(r, 2, val).font = data_font
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 30

    # ── Wire List ──
    ws2 = wb.create_sheet("Wire List")
    wire_hdrs = [
        "Wire #", "Signal Name", "Wire Type", "Gauge", "Color",
        "From Pin", "From Signal", "To Pin", "To Signal",
        "Length (m)", "Termination From", "Termination To", "Notes",
    ]
    _hdr(ws2, wire_hdrs)
    wires = db.query(Wire).filter(Wire.harness_id == harness_id).order_by(Wire.wire_number).all()
    for idx, w in enumerate(wires, 2):
        fp = db.query(Pin).filter(Pin.id == w.from_pin_id).first()
        tp = db.query(Pin).filter(Pin.id == w.to_pin_id).first()
        _row(ws2, idx, [
            w.wire_number, w.signal_name, _ev(w.wire_type), _ev(w.wire_gauge),
            w.wire_color_primary,
            fp.pin_number if fp else "", fp.signal_name if fp else "",
            tp.pin_number if tp else "", tp.signal_name if tp else "",
            w.length_m, w.termination_from, w.termination_to, w.notes,
        ])

    # ── Pin-to-Pin ──
    ws3 = wb.create_sheet("Pin-to-Pin")
    p2p_hdrs = [
        "From Unit", "From Conn", "From Pin", "From Signal",
        "Wire #", "Signal Name",
        "To Pin", "To Signal", "To Conn", "To Unit",
    ]
    _hdr(ws3, p2p_hdrs)
    for idx, w in enumerate(wires, 2):
        fp = db.query(Pin).filter(Pin.id == w.from_pin_id).first()
        tp = db.query(Pin).filter(Pin.id == w.to_pin_id).first()
        _row(ws3, idx, [
            fu.designation if fu else "", fc.designator if fc else "",
            fp.pin_number if fp else "", fp.signal_name if fp else "",
            w.wire_number, w.signal_name,
            tp.pin_number if tp else "", tp.signal_name if tp else "",
            tc.designator if tc else "", tu.designation if tu else "",
        ])

    # Auto-fit
    for s in wb.worksheets:
        for col in s.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            s.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=harness_{harness.harness_id or harness_id}.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  EXPORT — All Wiring
# ══════════════════════════════════════════════════════════════

@router.get("/export/all-wiring")
def export_all_wiring(
    project_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Master wiring export: all wires, unit connections, signal dictionary."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    wb = Workbook()
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    def _hdr(ws, hdrs):
        for i, h in enumerate(hdrs, 1):
            c = ws.cell(1, i, h)
            c.font = hdr_font; c.fill = hdr_fill; c.border = thin
        ws.freeze_panes = "A2"

    def _row(ws, r, vals):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = data_font; c.border = thin

    # ── Master Wire List ──
    ws1 = wb.active
    ws1.title = "Master Wire List"
    _hdr(ws1, [
        "Harness", "Wire #", "Signal", "Type", "Gauge", "Color",
        "From Unit", "From Conn", "From Pin",
        "To Unit", "To Conn", "To Pin", "Length (m)",
    ])
    harnesses = db.query(WireHarness).filter(WireHarness.project_id == project_id).all()
    row_idx = 2
    for h in harnesses:
        fu = db.query(Unit).filter(Unit.id == h.from_unit_id).first()
        fc = db.query(Connector).filter(Connector.id == h.from_connector_id).first()
        tu = db.query(Unit).filter(Unit.id == h.to_unit_id).first()
        tc = db.query(Connector).filter(Connector.id == h.to_connector_id).first()
        wires = db.query(Wire).filter(Wire.harness_id == h.id).order_by(Wire.wire_number).all()
        for w in wires:
            fp = db.query(Pin).filter(Pin.id == w.from_pin_id).first()
            tp = db.query(Pin).filter(Pin.id == w.to_pin_id).first()
            _row(ws1, row_idx, [
                h.name, w.wire_number, w.signal_name, _ev(w.wire_type), _ev(w.wire_gauge),
                w.wire_color_primary,
                fu.designation if fu else "", fc.designator if fc else "",
                fp.pin_number if fp else "",
                tu.designation if tu else "", tc.designator if tc else "",
                tp.pin_number if tp else "", w.length_m,
            ])
            row_idx += 1

    # ── Unit Connection Summary ──
    ws2 = wb.create_sheet("Unit Connections")
    _hdr(ws2, ["From Unit", "To Unit", "Harness", "Wire Count"])
    row_idx = 2
    for h in harnesses:
        fu = db.query(Unit).filter(Unit.id == h.from_unit_id).first()
        tu = db.query(Unit).filter(Unit.id == h.to_unit_id).first()
        wc = db.query(func.count(Wire.id)).filter(Wire.harness_id == h.id).scalar()
        _row(ws2, row_idx, [
            fu.designation if fu else "", tu.designation if tu else "", h.name, wc,
        ])
        row_idx += 1

    # ── Signal Dictionary ──
    ws3 = wb.create_sheet("Signal Dictionary")
    _hdr(ws3, ["Signal Name", "Unit", "Connector", "Pin", "Type", "Direction"])
    all_pins = (
        db.query(Pin).join(Connector).join(Unit)
        .filter(Unit.project_id == project_id)
        .order_by(Pin.signal_name, Unit.designation)
        .all()
    )
    row_idx = 2
    for p in all_pins:
        c = db.query(Connector).filter(Connector.id == p.connector_id).first()
        u = db.query(Unit).filter(Unit.id == c.unit_id).first() if c else None
        _row(ws3, row_idx, [
            p.signal_name, u.designation if u else "", c.designator if c else "",
            p.pin_number, _ev(p.signal_type), _ev(p.direction),
        ])
        row_idx += 1

    for s in wb.worksheets:
        for col in s.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            s.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=astra_{project.code}_wiring.xlsx"},
    )


# ══════════════════════════════════════════════════════════════
#  EXPORT — ICD Data Package (9 sheets)
# ══════════════════════════════════════════════════════════════

@router.get("/export/icd-data")
def export_icd_data(
    project_id: int = Query(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Complete ICD data package — 9 sheets covering all interface data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    wb = Workbook()
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    thin = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    def _hdr(ws, hdrs):
        for i, h in enumerate(hdrs, 1):
            c = ws.cell(1, i, h)
            c.font = hdr_font; c.fill = hdr_fill; c.border = thin
        ws.freeze_panes = "A2"

    def _row(ws, r, vals):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = data_font; c.border = thin

    systems = db.query(System).filter(System.project_id == project_id).order_by(System.name).all()
    units = db.query(Unit).filter(Unit.project_id == project_id).order_by(Unit.designation).all()

    # ── 1. Interface Summary (N² style) ──
    ws1 = wb.active
    ws1.title = "Interface Summary"
    interfaces = db.query(Interface).filter(Interface.project_id == project_id).all()
    _hdr(ws1, ["Interface ID", "Name", "Type", "Direction", "Source System",
                "Target System", "Criticality", "Status", "ICD Doc #"])
    for idx, iface in enumerate(interfaces, 2):
        ss = next((s for s in systems if s.id == iface.source_system_id), None)
        ts = next((s for s in systems if s.id == iface.target_system_id), None)
        _row(ws1, idx, [
            iface.interface_id, iface.name, _ev(iface.interface_type),
            _ev(iface.direction), ss.name if ss else "", ts.name if ts else "",
            _ev(iface.criticality), _ev(iface.status), iface.icd_document_number,
        ])

    # ── 2. Unit Catalog ──
    ws2 = wb.create_sheet("Unit Catalog")
    _hdr(ws2, ["Unit ID", "Designation", "Name", "System", "Type", "Status",
                "Part #", "Manufacturer", "Mass (kg)", "Power (W)", "Temp Min", "Temp Max"])
    for idx, u in enumerate(units, 2):
        s = next((s for s in systems if s.id == u.system_id), None)
        _row(ws2, idx, [
            u.unit_id, u.designation, u.name, s.name if s else "",
            _ev(u.unit_type), _ev(u.status), u.part_number, u.manufacturer,
            u.mass_kg, u.power_watts_nominal, u.temp_operating_min_c, u.temp_operating_max_c,
        ])

    # ── 3. Connector Pinouts ──
    ws3 = wb.create_sheet("Connector Pinouts")
    _hdr(ws3, ["Unit", "Connector", "Type", "Gender", "Pin #", "Signal", "Type", "Direction"])
    r = 2
    for u in units:
        for c in db.query(Connector).filter(Connector.unit_id == u.id).order_by(Connector.designator).all():
            for p in db.query(Pin).filter(Pin.connector_id == c.id).order_by(Pin.pin_number).all():
                _row(ws3, r, [u.designation, c.designator, _ev(c.connector_type),
                              _ev(c.gender), p.pin_number, p.signal_name,
                              _ev(p.signal_type), _ev(p.direction)])
                r += 1

    # ── 4. Bus Configuration ──
    ws4 = wb.create_sheet("Bus Configuration")
    _hdr(ws4, ["Unit", "Bus Name", "Protocol", "Role", "Address", "Data Rate",
                "Word Size", "Network", "Redundancy"])
    r = 2
    for u in units:
        for bd in db.query(BusDefinition).filter(BusDefinition.unit_id == u.id).all():
            _row(ws4, r, [u.designation, bd.name, _ev(bd.protocol), _ev(bd.bus_role),
                          bd.bus_address, bd.data_rate, bd.word_size_bits,
                          bd.bus_name_network, _ev(bd.redundancy)])
            r += 1

    # ── 5. Message Catalog ──
    ws5 = wb.create_sheet("Message Catalog")
    _hdr(ws5, ["Unit", "Bus", "Message", "Direction", "SA", "Words", "Rate (Hz)",
                "Field", "Type", "Word#", "Bits", "UoM", "Scale", "Min", "Max"])
    r = 2
    for u in units:
        for bd in db.query(BusDefinition).filter(BusDefinition.unit_id == u.id).all():
            for msg in db.query(MessageDefinition).filter(MessageDefinition.bus_def_id == bd.id).all():
                for f in db.query(MessageField).filter(MessageField.message_id == msg.id).order_by(MessageField.field_order).all():
                    _row(ws5, r, [u.designation, bd.name, msg.label, _ev(msg.direction),
                                  msg.subaddress, msg.word_count, msg.rate_hz,
                                  f.field_name, _ev(f.data_type), f.word_number,
                                  f.bit_length, f.unit_of_measure, f.scale_factor,
                                  f.min_value, f.max_value])
                    r += 1

    # ── 6. Wire Harnesses ──
    ws6 = wb.create_sheet("Wire Harnesses")
    _hdr(ws6, ["Harness", "From Unit", "From Conn", "To Unit", "To Conn",
                "Wire #", "Signal", "Type", "Gauge", "From Pin", "To Pin"])
    r = 2
    for h in db.query(WireHarness).filter(WireHarness.project_id == project_id).all():
        fu = db.query(Unit).filter(Unit.id == h.from_unit_id).first()
        fc = db.query(Connector).filter(Connector.id == h.from_connector_id).first()
        tu = db.query(Unit).filter(Unit.id == h.to_unit_id).first()
        tc = db.query(Connector).filter(Connector.id == h.to_connector_id).first()
        for w in db.query(Wire).filter(Wire.harness_id == h.id).order_by(Wire.wire_number).all():
            fp = db.query(Pin).filter(Pin.id == w.from_pin_id).first()
            tp = db.query(Pin).filter(Pin.id == w.to_pin_id).first()
            _row(ws6, r, [
                h.name, fu.designation if fu else "", fc.designator if fc else "",
                tu.designation if tu else "", tc.designator if tc else "",
                w.wire_number, w.signal_name, _ev(w.wire_type), _ev(w.wire_gauge),
                fp.pin_number if fp else "", tp.pin_number if tp else "",
            ])
            r += 1

    # ── 7. Signal Dictionary ──
    ws7 = wb.create_sheet("Signal Dictionary")
    _hdr(ws7, ["Signal Name", "Unit", "Connector", "Pin", "Type", "Direction"])
    all_pins = (
        db.query(Pin).join(Connector).join(Unit)
        .filter(Unit.project_id == project_id)
        .order_by(Pin.signal_name)
        .all()
    )
    r = 2
    for p in all_pins:
        c = db.query(Connector).filter(Connector.id == p.connector_id).first()
        u = db.query(Unit).filter(Unit.id == c.unit_id).first() if c else None
        _row(ws7, r, [p.signal_name, u.designation if u else "",
                       c.designator if c else "", p.pin_number,
                       _ev(p.signal_type), _ev(p.direction)])
        r += 1

    # ── 8. Environmental Summary ──
    ws8 = wb.create_sheet("Environmental Summary")
    _hdr(ws8, ["Unit", "Category", "Standard", "Method", "Level",
                "Value", "Unit", "Min", "Max", "Status"])
    r = 2
    for u in units:
        for es in db.query(UnitEnvironmentalSpec).filter(UnitEnvironmentalSpec.unit_id == u.id).all():
            _row(ws8, r, [u.designation, _ev(es.category), _ev(es.standard),
                          es.test_method, es.test_level, es.limit_value,
                          es.limit_unit, es.limit_min, es.limit_max,
                          es.compliance_status])
            r += 1

    # ── 9. Requirements Trace ──
    ws9 = wb.create_sheet("Requirements Trace")
    _hdr(ws9, ["Entity Type", "Entity ID", "Link Type", "Requirement ID",
                "Requirement Title", "Auto-Generated", "Status"])
    r = 2
    links = db.query(InterfaceRequirementLink).all()
    for lk in links:
        req = db.query(Requirement).filter(Requirement.id == lk.requirement_id).first()
        _row(ws9, r, [
            _ev(lk.entity_type), lk.entity_id, _ev(lk.link_type),
            req.req_id if req else "", req.title if req else "",
            lk.auto_generated, _ev(lk.status),
        ])
        r += 1

    # Auto-fit all sheets
    for s in wb.worksheets:
        for col in s.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            s.column_dimensions[get_column_letter(col[0].column)].width = min(ml + 3, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=astra_{project.code}_icd_data.xlsx"},
    )

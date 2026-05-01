"""
ASTRA — Requirements Import Router
=====================================
File: backend/app/routers/imports.py

Endpoints:
  POST /imports/requirements          — upload CSV/XLSX, returns preview with quality scores
  POST /imports/requirements/confirm   — confirm import after preview approval
  GET  /imports/template               — download a blank CSV template
"""

import io
import csv
import logging
from typing import Optional, List
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.database import get_db
from app.dependencies.project_access import _check_membership
from app.models import Requirement, Project, User, RequirementHistory
from app.schemas import RequirementResponse
from app.services.auth import get_current_user
from app.services.quality_checker import check_requirement_quality, generate_requirement_id
from app.services.security.spreadsheet import (
    assert_workbook_size_ok, sanitize_filename, validate_upload,
)

try:
    from app.services.rbac import require_permission
except ImportError:
    def require_permission(action):
        return get_current_user

try:
    from app.services.audit_service import record_event as _audit
except ImportError:
    def _audit(*a, **kw):
        pass

logger = logging.getLogger("astra.imports")

router = APIRouter(prefix="/imports", tags=["Import"])

# ── Valid values ──

VALID_TYPES = {
    "functional", "performance", "interface", "environmental",
    "constraint", "safety", "security", "reliability",
    "maintainability", "derived",
}
VALID_PRIORITIES = {"critical", "high", "medium", "low"}
VALID_LEVELS = {"L1", "L2", "L3", "L4", "L5"}

# ── Column aliases (flexible header matching) ──

COLUMN_ALIASES = {
    "title": ["title", "name", "requirement_title", "req_title"],
    "statement": ["statement", "description", "requirement", "shall_statement", "text", "body"],
    "rationale": ["rationale", "reason", "justification", "why"],
    "req_type": ["req_type", "type", "requirement_type", "category"],
    "priority": ["priority", "pri", "importance"],
    "level": ["level", "lvl", "hierarchy", "tier"],
    "parent_req_id": ["parent_req_id", "parent", "parent_id", "parent_requirement"],
}


def _match_column(header: str) -> Optional[str]:
    """Match a CSV/XLSX header to a known field name."""
    h = header.strip().lower().replace(" ", "_").replace("-", "_")
    for field, aliases in COLUMN_ALIASES.items():
        if h in aliases:
            return field
    return None


# ── Schemas ──

class ImportRowPreview(BaseModel):
    row_number: int
    title: str = ""
    statement: str = ""
    rationale: str = ""
    req_type: str = "functional"
    priority: str = "medium"
    level: str = "L1"
    parent_req_id: str = ""
    quality_score: float = 0.0
    quality_passed: bool = False
    warnings: List[str] = Field(default_factory=list)
    errors: List[str] = Field(default_factory=list)
    included: bool = True


class ImportPreviewResponse(BaseModel):
    filename: str
    total_rows: int
    valid_rows: int
    error_rows: int
    column_mapping: dict
    rows: List[ImportRowPreview]


class ImportConfirmRequest(BaseModel):
    project_id: int
    rows: List[ImportRowPreview]


class ImportConfirmResponse(BaseModel):
    created: int
    skipped: int
    errors: List[str]
    requirements: List[dict] = Field(default_factory=list)


# ── Parse helpers ──

def _parse_csv(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse CSV bytes into headers + rows."""
    text = content.decode("utf-8-sig")  # Handle BOM
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    rows = list(reader)
    return headers, rows


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict]]:
    """Parse XLSX bytes into headers + rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(400, "openpyxl not installed — cannot parse XLSX files")

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    # F-018 layer 4: cap sheet count + per-sheet row count. Raises 413.
    assert_workbook_size_ok(wb)
    ws = wb.active
    if ws is None:
        return [], []

    row_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        return [], []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(header_row)]
    rows = []
    for row in row_iter:
        if all(c is None for c in row):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = str(val).strip() if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return headers, rows


def _validate_row(row: dict, mapping: dict, row_num: int,
                  existing_req_ids: set) -> ImportRowPreview:
    """Validate a single parsed row and return a preview object."""
    errors: list[str] = []
    warnings: list[str] = []

    # Extract mapped values
    title = ""
    statement = ""
    rationale = ""
    req_type = "functional"
    priority = "medium"
    level = "L1"
    parent_req_id = ""

    for orig_col, mapped_field in mapping.items():
        val = row.get(orig_col, "").strip()
        if not val:
            continue
        if mapped_field == "title":
            title = val
        elif mapped_field == "statement":
            statement = val
        elif mapped_field == "rationale":
            rationale = val
        elif mapped_field == "req_type":
            req_type = val.lower()
        elif mapped_field == "priority":
            priority = val.lower()
        elif mapped_field == "level":
            level = val.upper()
        elif mapped_field == "parent_req_id":
            parent_req_id = val

    # Required fields
    if not title:
        errors.append("Missing title")
    if not statement:
        errors.append("Missing statement")

    # Validate enums
    if req_type and req_type not in VALID_TYPES:
        warnings.append(f"Unknown type '{req_type}' — defaulting to functional")
        req_type = "functional"
    if priority and priority not in VALID_PRIORITIES:
        warnings.append(f"Unknown priority '{priority}' — defaulting to medium")
        priority = "medium"
    if level and level not in VALID_LEVELS:
        warnings.append(f"Unknown level '{level}' — defaulting to L1")
        level = "L1"

    # Parent validation
    if parent_req_id and parent_req_id not in existing_req_ids:
        warnings.append(f"Parent '{parent_req_id}' not found in project — will be ignored")

    # Quality check
    quality_score = 0.0
    quality_passed = False
    if statement:
        qr = check_requirement_quality(statement, title, rationale)
        quality_score = qr["score"]
        quality_passed = qr["passed"]
        warnings.extend(qr.get("warnings", []))

    return ImportRowPreview(
        row_number=row_num,
        title=title,
        statement=statement,
        rationale=rationale,
        req_type=req_type,
        priority=priority,
        level=level,
        parent_req_id=parent_req_id,
        quality_score=quality_score,
        quality_passed=quality_passed,
        warnings=warnings,
        errors=errors,
        included=len(errors) == 0,
    )


# ══════════════════════════════════════
#  Preview endpoint
# ══════════════════════════════════════

@router.post("/requirements", response_model=ImportPreviewResponse)
async def preview_import(
    request: Request,
    project_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("requirements.create")),
):
    """
    Upload a CSV or XLSX file and get a preview of what will be imported.
    Each row is validated and quality-checked but NOT yet saved.

    AUDIT_FINDINGS F-018:
      - BodySizeLimitMiddleware (in main.py) rejects bodies above
        MAX_UPLOAD_BYTES with 413 before this handler runs.
      - Content-Type allowlist + magic-byte sniff via
        services.security.spreadsheet.validate_upload (415 on miss).
      - Sheet- and row-count caps via assert_workbook_size_ok (413 on
        overflow).
      - Filename sanitised before being echoed in the response.
    """
    # Validate project
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    # F-211: enforce membership so a user with `requirements.create`
    # cannot preview-import into a project they were never added to.
    _check_membership(db, project_id, current_user)

    # Read file
    content = await file.read()

    # F-018 layer 2-3: declared MIME + magic-byte sniff. Raises 415/413.
    sniffed_kind = validate_upload(
        content=content,
        declared_content_type=file.content_type,
        expected_kind="csv_or_xlsx",
    )

    # F-018 layer 5: sanitise filename before echo.
    filename = sanitize_filename(file.filename, default="upload")

    # Parse — sniffed_kind drives the parser, NOT the user's filename ext.
    if sniffed_kind == "csv":
        headers, rows = _parse_csv(content)
    else:  # "xlsx"
        headers, rows = _parse_xlsx(content)

    if not headers:
        raise HTTPException(400, "No headers found in file")
    if not rows:
        raise HTTPException(400, "No data rows found in file")

    # Auto-map columns
    mapping: dict[str, str] = {}
    for h in headers:
        matched = _match_column(h)
        if matched:
            mapping[h] = matched

    if "statement" not in mapping.values() and "title" not in mapping.values():
        raise HTTPException(400,
            f"Could not map any columns. Found headers: {headers}. "
            f"Expected at least 'title' and 'statement'.")

    # Get existing req_ids for parent matching
    existing = db.query(Requirement.req_id).filter(
        Requirement.project_id == project_id,
        Requirement.status != "deleted",
    ).all()
    existing_req_ids = {r[0] for r in existing}

    # Validate each row
    previews: list[ImportRowPreview] = []
    for i, row in enumerate(rows, start=2):  # Row 2 = first data row (1 = header)
        preview = _validate_row(row, mapping, i, existing_req_ids)
        previews.append(preview)

    valid = sum(1 for p in previews if p.included)
    errors = sum(1 for p in previews if not p.included)

    return ImportPreviewResponse(
        filename=filename,
        total_rows=len(previews),
        valid_rows=valid,
        error_rows=errors,
        column_mapping=mapping,
        rows=previews,
    )


# ══════════════════════════════════════
#  Confirm endpoint
# ══════════════════════════════════════

@router.post("/requirements/confirm", response_model=ImportConfirmResponse)
def confirm_import(
    data: ImportConfirmRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_permission("requirements.create")),
):
    """
    Create requirements from the approved preview rows.
    Only rows with `included=True` are imported.

    F-070: pass `request` through to `_audit` so the import.completed
    event carries the originating IP / User-Agent. Pre-fix the audit
    row had blank user_ip / user_agent because the dep chain didn't
    take a Request, so the audit middleware's request-scoped
    contextvar wasn't populated for this handler.
    """
    project = db.query(Project).filter(Project.id == data.project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    # F-211: enforce membership before writing N requirements + N
    # history rows + an audit row attributed to current_user.
    _check_membership(db, data.project_id, current_user)

    # Build parent lookup
    existing = db.query(Requirement).filter(
        Requirement.project_id == data.project_id,
        Requirement.status != "deleted",
    ).all()
    req_id_to_pk: dict[str, int] = {r.req_id: r.id for r in existing}

    created = 0
    skipped = 0
    errors: list[str] = []
    created_reqs: list[dict] = []

    for row in data.rows:
        if not row.included:
            skipped += 1
            continue

        if not row.title or not row.statement:
            errors.append(f"Row {row.row_number}: missing title or statement")
            skipped += 1
            continue

        # F-055: per-row SAVEPOINT so a failure inside this row's
        # add+flush+history block rolls back ONLY this row, not the
        # whole batch. Pre-fix the outer try/except left orphan
        # objects in the session that could taint subsequent rows.
        sp = db.begin_nested()
        try:
            # Generate req_id
            count = db.query(func.count(Requirement.id)).filter(
                Requirement.project_id == data.project_id,
                Requirement.req_type == row.req_type,
            ).scalar()
            req_id = generate_requirement_id(project.code, row.req_type, count + 1)

            # Resolve parent
            parent_id = None
            if row.parent_req_id and row.parent_req_id in req_id_to_pk:
                parent_id = req_id_to_pk[row.parent_req_id]

            # Quality check
            qr = check_requirement_quality(row.statement, row.title, row.rationale)

            req = Requirement(
                req_id=req_id,
                title=row.title,
                statement=row.statement,
                rationale=row.rationale or None,
                req_type=row.req_type,
                priority=row.priority,
                level=row.level,
                project_id=data.project_id,
                owner_id=current_user.id,
                created_by_id=current_user.id,
                parent_id=parent_id,
                quality_score=qr["score"],
            )
            db.add(req)
            db.flush()

            # Record history
            history = RequirementHistory(
                requirement_id=req.id,
                version=1,
                field_changed="created",
                old_value=None,
                new_value=req.req_id,
                change_description=f"Imported from file (row {row.row_number})",
                changed_by_id=current_user.id,
                changed_at=datetime.utcnow(),
            )
            db.add(history)
            db.flush()
            sp.commit()  # release savepoint — row's writes are durable on outer commit

            # Track for parent resolution of subsequent rows
            req_id_to_pk[req_id] = req.id

            created_reqs.append({
                "id": req.id,
                "req_id": req_id,
                "title": row.title,
                "quality_score": qr["score"],
            })
            created += 1

        except Exception as exc:
            sp.rollback()  # F-055: roll back this row only
            errors.append(f"Row {row.row_number}: {str(exc)}")
            skipped += 1

    db.commit()

    try:
        _audit(db, "import.completed", "project", data.project_id,
               current_user.id,
               {"created": created, "skipped": skipped, "errors": len(errors)},
               project_id=data.project_id, request=request)
    except Exception:
        pass

    return ImportConfirmResponse(
        created=created,
        skipped=skipped,
        errors=errors,
        requirements=created_reqs,
    )


# ══════════════════════════════════════
#  Template download
# ══════════════════════════════════════

TEMPLATE_CSV = (
    "title,statement,rationale,req_type,priority,level,parent_req_id\n"
    "User Authentication,"
    "The system shall authenticate users via username and password within 3 seconds.,"
    "Secure authentication protects sensitive data.,"
    "functional,high,L1,\n"
    "Encryption at Rest,"
    "The system shall encrypt all stored data using AES-256.,"
    "ITAR compliance requires data protection.,"
    "security,critical,L2,FR-001\n"
)


@router.get("/template")
def download_template(
    current_user: User = Depends(get_current_user),
):
    """Download a pre-formatted CSV template with example rows.

    F-072: requires auth. The template is innocuous (no project data,
    just the column-header schema) but unauthenticated download
    leaks the full requirements-import field set publicly. Cheap to
    gate, so we gate it.
    """
    return StreamingResponse(
        io.BytesIO(TEMPLATE_CSV.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="astra_import_template.csv"'},
    )

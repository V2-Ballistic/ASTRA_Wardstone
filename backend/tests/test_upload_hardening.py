"""
ASTRA — File-upload hardening tests (F-018)
============================================
File: backend/tests/test_upload_hardening.py

Layered defenses:
  1. BodySizeLimitMiddleware (Content-Length > MAX_UPLOAD_BYTES → 413)
  2. validate_upload (declared MIME + magic-byte sniff → 415)
  3. assert_workbook_size_ok (sheet/row caps → 413)
  4. formula_safe (cells starting = + - @ get a leading apostrophe)
  5. sanitize_filename (echoed names are basename-only, allowlisted chars)
"""

from __future__ import annotations

import io
import os
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.models import Project, User, UserRole
from app.models.project_member import ProjectMember
from app.services.auth import create_access_token, get_password_hash
from app.services.security.spreadsheet import (
    formula_safe, sanitize_filename, sniff_content_type, validate_upload,
)


# ──────────────────────────────────────
#  Pure-function unit tests
# ──────────────────────────────────────


class TestSniff:
    def test_xlsx_magic_bytes(self):
        assert sniff_content_type(b"PK\x03\x04rest") == "xlsx"

    def test_csv_printable(self):
        assert sniff_content_type(b"a,b,c\n1,2,3\n") == "csv"

    def test_binary_garbage(self):
        assert sniff_content_type(b"\x00\x01\x02\xff\xfe garbage") == "binary"

    def test_empty(self):
        assert sniff_content_type(b"") == "binary"


class TestFormulaSafe:
    @pytest.mark.parametrize("dangerous,expected", [
        ("=cmd|' /C calc'!A1", "'=cmd|' /C calc'!A1"),
        ("+1+1", "'+1+1"),
        ("-2", "'-2"),
        ("@SUM(1)", "'@SUM(1)"),
    ])
    def test_dangerous_prefix_escaped(self, dangerous, expected):
        assert formula_safe(dangerous) == expected

    @pytest.mark.parametrize("benign", [
        "hello world",
        "FR-001 The system shall…",
        "1.5",
        "",
    ])
    def test_benign_pass_through(self, benign):
        assert formula_safe(benign) == benign

    def test_none_becomes_empty(self):
        assert formula_safe(None) == ""

    def test_non_string_stringified(self):
        assert formula_safe(42) == "42"


class TestSanitizeFilename:
    @pytest.mark.parametrize("dangerous,expected", [
        ("../../../etc/passwd", "passwd"),
        ("..\\..\\Windows\\system32\\foo.exe", "foo.exe"),
        ("rm -rf $HOME", "rm_-rf__HOME"),
        ("normal.xlsx", "normal.xlsx"),
        (".hidden", "hidden"),
        ("", "upload"),
        (None, "upload"),
    ])
    def test_sanitize(self, dangerous, expected):
        assert sanitize_filename(dangerous) == expected


# ──────────────────────────────────────
#  Middleware: 413 on Content-Length > MAX_UPLOAD_BYTES
# ──────────────────────────────────────


def _auth_headers_for(db_session, role="project_manager"):
    u = User(
        username="uploader", email="uploader@example.com",
        hashed_password=get_password_hash("UploaderPass1"),
        full_name="Uploader",
        role=role, department="Eng", is_active=True,
    )
    db_session.add(u); db_session.commit(); db_session.refresh(u)
    return {"Authorization": f"Bearer {create_access_token(data={'sub': u.username})}"}, u


def _project_with_member(db_session, owner: User, code="UP1"):
    p = Project(code=code, name="Upload P", owner_id=owner.id, status="active")
    db_session.add(p); db_session.commit(); db_session.refresh(p)
    db_session.add(ProjectMember(project_id=p.id, user_id=owner.id, added_by_id=owner.id))
    db_session.commit()
    return p


class TestBodySizeMiddleware:
    def test_oversize_content_length_returns_413_before_handler(
        self, client, db_session,
    ):
        """
        Middleware fires on Content-Length alone — we don't even need to
        send the body. Server should reject with 413 immediately.
        """
        from app.middleware.body_size_limit import _DEFAULT_MAX_BYTES

        headers, _ = _auth_headers_for(db_session)
        oversize = _DEFAULT_MAX_BYTES + 1
        # We send a tiny body but lie in Content-Length to fire the
        # middleware's pre-handler check. (httpx will set the actual
        # Content-Length from the data; we override via headers.)
        r = client.post(
            "/api/v1/imports/requirements?project_id=1",
            content=b"x",
            headers={
                **headers,
                "Content-Length": str(oversize),
            },
        )
        # Middleware short-circuits before the route runs → 413 with our detail.
        assert r.status_code == 413, r.text
        assert "limit" in r.json()["detail"].lower()


# ──────────────────────────────────────
#  Endpoint-level: 415 on Content-Type mismatch
# ──────────────────────────────────────


class TestPreviewImportContentType:
    def test_csv_declared_as_xlsx_returns_415(self, client, db_session):
        headers, owner = _auth_headers_for(db_session)
        project = _project_with_member(db_session, owner, code="UP2")

        csv_body = b"title,statement\nFoo,The system shall do X.\n"
        r = client.post(
            f"/api/v1/imports/requirements?project_id={project.id}",
            files={"file": ("file.csv", csv_body, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=headers,
        )
        assert r.status_code == 415, r.text

    def test_garbage_binary_declared_as_csv_returns_415(self, client, db_session):
        headers, owner = _auth_headers_for(db_session)
        project = _project_with_member(db_session, owner, code="UP3")

        garbage = b"\x00\x01\x02\xff\xfe random garbage \x00"
        r = client.post(
            f"/api/v1/imports/requirements?project_id={project.id}",
            files={"file": ("file.csv", garbage, "text/csv")},
            headers=headers,
        )
        assert r.status_code == 415, r.text

    def test_real_csv_passes_validation(self, client, db_session):
        """Sanity check: a well-formed CSV gets to (and through) the handler."""
        headers, owner = _auth_headers_for(db_session)
        project = _project_with_member(db_session, owner, code="UP4")

        csv_body = (
            b"title,statement\n"
            b"Login,The system shall authenticate users within 2 seconds.\n"
        )
        r = client.post(
            f"/api/v1/imports/requirements?project_id={project.id}",
            files={"file": ("file.csv", csv_body, "text/csv")},
            headers=headers,
        )
        assert r.status_code == 200, r.text


# ──────────────────────────────────────
#  Workbook size caps: 413 after parse
# ──────────────────────────────────────


def _build_xlsx(rows: int) -> bytes:
    """Build an in-memory XLSX with N data rows + 1 header."""
    wb = Workbook(write_only=False)
    ws = wb.active
    ws.append(["title", "statement"])
    for i in range(rows):
        ws.append([f"Req {i}", f"The system shall do thing {i} within 5 seconds."])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestWorkbookSizeCaps:
    def test_xlsx_above_row_cap_returns_413(self, client, db_session, monkeypatch):
        """Override MAX_UPLOAD_ROWS to a small number and POST one row above it."""
        monkeypatch.setattr(
            "app.services.security.spreadsheet.MAX_ROWS", 5,
        )
        headers, owner = _auth_headers_for(db_session)
        project = _project_with_member(db_session, owner, code="UP5")

        body = _build_xlsx(rows=10)  # > 5
        r = client.post(
            f"/api/v1/imports/requirements?project_id={project.id}",
            files={"file": (
                "big.xlsx", body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=headers,
        )
        assert r.status_code == 413, r.text
        assert "rows" in r.json()["detail"].lower()

    def test_xlsx_above_sheet_cap_returns_413(self, client, db_session, monkeypatch):
        monkeypatch.setattr(
            "app.services.security.spreadsheet.MAX_SHEETS", 2,
        )
        headers, owner = _auth_headers_for(db_session)
        project = _project_with_member(db_session, owner, code="UP6")

        wb = Workbook()
        wb.active.title = "Sheet1"
        wb.active.append(["title", "statement"])
        wb.create_sheet("Sheet2")
        wb.create_sheet("Sheet3")
        buf = io.BytesIO(); wb.save(buf)
        body = buf.getvalue()

        r = client.post(
            f"/api/v1/imports/requirements?project_id={project.id}",
            files={"file": (
                "many.xlsx", body,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )},
            headers=headers,
        )
        assert r.status_code == 413, r.text
        assert "sheets" in r.json()["detail"].lower()

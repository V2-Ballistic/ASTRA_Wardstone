"""
ASTRA — Persistent report jobs (F-019 + F-032)
================================================
File: backend/tests/test_report_jobs.py

Covers:
  * The synchronous report endpoints now write a COMPLETED row to
    ``report_jobs`` (replacing the old in-memory ``_report_history``).
  * ``GET /reports/history`` reads from the DB and is project-scoped.
  * The async pattern (POST → poll → download) round-trips correctly,
    including cross-project access denial and unknown-type handling.
"""

from __future__ import annotations

import io

import pytest
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook

from app.models import Project, User
from app.models.project_member import ProjectMember
from app.models.report_job import ReportJob, ReportJobStatus
from app.services.auth import create_access_token, get_password_hash


# ──────────────────────────────────────
#  Helpers
# ──────────────────────────────────────


def _make_user(db_session, *, username: str, role: str = "developer") -> User:
    u = User(
        username=username,
        email=f"{username}@example.com",
        hashed_password=get_password_hash("ReportPass1"),
        full_name=username.title(),
        role=role,
        department="Eng",
        is_active=True,
    )
    db_session.add(u)
    db_session.commit()
    db_session.refresh(u)
    return u


def _headers(user: User) -> dict:
    return {"Authorization": f"Bearer {create_access_token(data={'sub': user.username})}"}


def _make_project(db_session, owner: User, code: str) -> Project:
    p = Project(code=code, name=f"P {code}", owner_id=owner.id, status="active")
    db_session.add(p)
    db_session.commit()
    db_session.refresh(p)
    db_session.add(ProjectMember(project_id=p.id, user_id=owner.id, added_by_id=owner.id))
    db_session.commit()
    return p


def _patch_session_factory(monkeypatch, db_session) -> None:
    """
    BackgroundTasks open their own SessionLocal, but the test fixture
    binds the session to a per-test sqlite engine. Rebind the factory
    used inside ``run_job`` so background work hits the same engine
    rows the test is asserting against.
    """
    test_factory = sessionmaker(bind=db_session.bind)
    monkeypatch.setattr(
        "app.services.reports.jobs.SessionLocal", test_factory,
    )


# ──────────────────────────────────────
#  Synchronous path persists into report_jobs
# ──────────────────────────────────────


class TestSyncPersistsHistory:
    def test_quality_report_writes_completed_row(self, client, db_session):
        owner = _make_user(db_session, username="rep_owner")
        project = _make_project(db_session, owner, "RPT1")

        r = client.get(
            f"/api/v1/reports/quality?project_id={project.id}&format=xlsx",
            headers=_headers(owner),
        )
        assert r.status_code == 200, r.text
        assert "Content-Disposition" in r.headers
        assert r.content[:2] == b"PK"  # xlsx zip magic

        rows = (
            db_session.query(ReportJob)
            .filter(ReportJob.project_id == project.id)
            .all()
        )
        assert len(rows) == 1
        row = rows[0]
        assert row.status == ReportJobStatus.COMPLETED
        assert row.report_type == "quality"
        assert row.format == "xlsx"
        assert row.requested_by_id == owner.id
        assert row.result_filename
        assert row.result_blob and len(row.result_blob) > 0
        assert row.completed_at is not None


# ──────────────────────────────────────
#  /reports/history is DB-backed and scoped
# ──────────────────────────────────────


class TestHistoryEndpoint:
    def test_history_requires_project_id(self, client, db_session):
        owner = _make_user(db_session, username="rep_owner2")
        # Without project_id — FastAPI returns 422 on missing required query.
        r = client.get("/api/v1/reports/history", headers=_headers(owner))
        assert r.status_code == 422, r.text

    def test_history_scoped_by_membership(self, client, db_session):
        owner = _make_user(db_session, username="rep_a")
        outsider = _make_user(db_session, username="rep_b")
        project = _make_project(db_session, owner, "RPT2")

        # Owner generates a report.
        r = client.get(
            f"/api/v1/reports/quality?project_id={project.id}",
            headers=_headers(owner),
        )
        assert r.status_code == 200

        # Outsider sees 403, not the row.
        r2 = client.get(
            f"/api/v1/reports/history?project_id={project.id}",
            headers=_headers(outsider),
        )
        assert r2.status_code == 403, r2.text

    def test_history_returns_persisted_rows(self, client, db_session):
        owner = _make_user(db_session, username="rep_c")
        project = _make_project(db_session, owner, "RPT3")

        for _ in range(3):
            client.get(
                f"/api/v1/reports/quality?project_id={project.id}",
                headers=_headers(owner),
            )

        r = client.get(
            f"/api/v1/reports/history?project_id={project.id}",
            headers=_headers(owner),
        )
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["total"] == 3
        assert len(body["items"]) == 3
        for item in body["items"]:
            assert item["report_type"] == "quality"
            assert item["status"] == "completed"
            assert item["project_id"] == project.id


# ──────────────────────────────────────
#  Async POST → poll → download
# ──────────────────────────────────────


class TestAsyncJobFlow:
    def test_post_returns_job_id_and_completes(
        self, client, db_session, monkeypatch,
    ):
        _patch_session_factory(monkeypatch, db_session)

        owner = _make_user(db_session, username="async_owner")
        project = _make_project(db_session, owner, "RPT4")

        r = client.post(
            f"/api/v1/reports/quality/jobs?project_id={project.id}&format=xlsx",
            headers=_headers(owner),
        )
        assert r.status_code == 202, r.text
        job_id = r.json()["job_id"]
        assert isinstance(job_id, int)

        # BackgroundTask runs after the response — by the time the test
        # continues, the row should be COMPLETED.
        r2 = client.get(
            f"/api/v1/reports/jobs/{job_id}",
            headers=_headers(owner),
        )
        assert r2.status_code == 200, r2.text
        assert r2.json()["status"] == "completed"
        assert r2.json()["filename"]

        r3 = client.get(
            f"/api/v1/reports/jobs/{job_id}/download",
            headers=_headers(owner),
        )
        assert r3.status_code == 200, r3.text
        assert r3.content[:2] == b"PK"

        # The downloaded blob is a real xlsx — openpyxl can parse it.
        wb = load_workbook(io.BytesIO(r3.content), read_only=True)
        assert wb.sheetnames

    def test_unknown_report_type_returns_404(self, client, db_session):
        owner = _make_user(db_session, username="async_unknown")
        project = _make_project(db_session, owner, "RPT5")

        r = client.post(
            f"/api/v1/reports/no-such-report/jobs?project_id={project.id}",
            headers=_headers(owner),
        )
        assert r.status_code == 404, r.text

    def test_non_member_cannot_enqueue(self, client, db_session):
        owner = _make_user(db_session, username="async_owner2")
        outsider = _make_user(db_session, username="async_outsider")
        project = _make_project(db_session, owner, "RPT6")

        r = client.post(
            f"/api/v1/reports/quality/jobs?project_id={project.id}",
            headers=_headers(outsider),
        )
        assert r.status_code == 403, r.text

    def test_non_member_cannot_view_or_download_job(
        self, client, db_session, monkeypatch,
    ):
        _patch_session_factory(monkeypatch, db_session)

        owner = _make_user(db_session, username="async_owner3")
        outsider = _make_user(db_session, username="async_outsider2")
        project = _make_project(db_session, owner, "RPT7")

        r = client.post(
            f"/api/v1/reports/quality/jobs?project_id={project.id}",
            headers=_headers(owner),
        )
        assert r.status_code == 202, r.text
        job_id = r.json()["job_id"]

        r2 = client.get(
            f"/api/v1/reports/jobs/{job_id}",
            headers=_headers(outsider),
        )
        assert r2.status_code == 403, r2.text

        r3 = client.get(
            f"/api/v1/reports/jobs/{job_id}/download",
            headers=_headers(outsider),
        )
        assert r3.status_code == 403, r3.text

    def test_get_job_404_for_unknown_id(self, client, db_session):
        owner = _make_user(db_session, username="async_404")
        # Make a project so the user has at least one membership context.
        _make_project(db_session, owner, "RPT8")

        r = client.get("/api/v1/reports/jobs/999999", headers=_headers(owner))
        assert r.status_code == 404, r.text


# ──────────────────────────────────────
#  Failure path: report generator raises → job → FAILED
# ──────────────────────────────────────


class TestAsyncFailureMarksJobFailed:
    def test_generator_exception_marks_failed(
        self, client, db_session, monkeypatch,
    ):
        _patch_session_factory(monkeypatch, db_session)

        # Replace the registered generator with one that raises.
        from app.services.reports import REPORT_REGISTRY
        from app.services.reports.base import ReportGenerator

        class BoomReport(ReportGenerator):
            name = "boom"
            supported_formats = ["xlsx"]

            def generate(self, project_id, db, options=None):
                raise RuntimeError("boom")

        monkeypatch.setitem(REPORT_REGISTRY, "quality", BoomReport)

        owner = _make_user(db_session, username="boom_owner")
        project = _make_project(db_session, owner, "RPT9")

        r = client.post(
            f"/api/v1/reports/quality/jobs?project_id={project.id}",
            headers=_headers(owner),
        )
        assert r.status_code == 202, r.text
        job_id = r.json()["job_id"]

        r2 = client.get(
            f"/api/v1/reports/jobs/{job_id}",
            headers=_headers(owner),
        )
        assert r2.status_code == 200, r2.text
        body = r2.json()
        assert body["status"] == "failed"
        assert body["error"] and "boom" in body["error"]

        # Download on a non-completed job is 409.
        r3 = client.get(
            f"/api/v1/reports/jobs/{job_id}/download",
            headers=_headers(owner),
        )
        assert r3.status_code == 409, r3.text

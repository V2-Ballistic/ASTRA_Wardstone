"""
ASTRA — Interface Module Tests
=================================
File: backend/tests/test_interface.py
Path: C:\\Users\\Mason\\Documents\\ASTRA\\backend\\tests\\test_interface.py

56 tests across 10 classes:
  TestSystemCRUD (5)
  TestUnitCRUD (8)
  TestConnectorAndPins (8)
  TestBusDefinition (5)
  TestMessageDefinition (5)
  TestWireHarness (6)
  TestAutoWire (3)
  TestAutoRequirementGeneration (8) — CRITICAL
  TestExcelImportExport (5)
  TestSignalTrace (3)
"""

import io
import pytest
from app.models.interface import (
    System, Unit, Connector, Pin, BusDefinition,
    PinBusAssignment, MessageDefinition, MessageField,
    WireHarness, Wire, Interface,
    UnitEnvironmentalSpec, InterfaceRequirementLink,
)
from app.models import Requirement, Verification


# ══════════════════════════════════════════════════════════════
#  Fixtures
# ══════════════════════════════════════════════════════════════

@pytest.fixture()
def test_system(db_session, test_user, test_project):
    s = System(
        system_id="SYS-001", name="Radar Subsystem", abbreviation="RADAR",
        system_type="subsystem", project_id=test_project.id, owner_id=test_user.id,
    )
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture()
def test_system_b(db_session, test_user, test_project):
    s = System(
        system_id="SYS-002", name="C2 Subsystem", abbreviation="C2",
        system_type="subsystem", project_id=test_project.id, owner_id=test_user.id,
    )
    db_session.add(s)
    db_session.commit()
    db_session.refresh(s)
    return s


@pytest.fixture()
def test_unit(db_session, test_project, test_system):
    u = Unit(
        unit_id="UNIT-001", name="Radar Signal Processor", designation="RSP-100",
        part_number="RSP-100-A1", manufacturer="Raytheon", unit_type="processor",
        status="concept", system_id=test_system.id, project_id=test_project.id,
        temp_operating_min_c=-40.0, temp_operating_max_c=85.0,
        vibration_random_grms=7.5, shock_mechanical_g=40.0,
        emi_ce102_limit_dbua=60.0, emi_rs103_limit_vm=20.0,
        esd_hbm_v=2000.0, mass_kg=2.3, power_watts_nominal=45.0,
    )
    db_session.add(u)
    db_session.commit()
    db_session.refresh(u)
    return u


@pytest.fixture()
def test_unit_b(db_session, test_project, test_system_b):
    u = Unit(
        unit_id="UNIT-002", name="C2 Processor", designation="C2P-200",
        part_number="C2P-200-B1", manufacturer="BAE", unit_type="processor",
        status="concept", system_id=test_system_b.id, project_id=test_project.id,
    )
    db_session.add(u)
    db_session.commit()
    db_session.refresh(u)
    return u


@pytest.fixture()
def test_connector_a(db_session, test_project, test_unit):
    """Connector J1 on RSP-100 with 5 pins."""
    c = Connector(
        connector_id="CONN-001", designator="J1", name="1553 Bus A",
        connector_type="mil_dtl_38999_series_iii", gender="female_socket",
        total_contacts=25, unit_id=test_unit.id, project_id=test_project.id,
    )
    db_session.add(c)
    db_session.flush()
    pin_defs = [
        ("1", "28V_PWR", "power_primary", "power_source"),
        ("2", "28V_RTN", "power_return", "power_return"),
        ("3", "1553A_HI", "signal_digital_differential", "bidirectional"),
        ("4", "1553A_LO", "signal_digital_differential", "bidirectional"),
        ("5", "SPARE_1", "spare", "no_connect"),
    ]
    pins = []
    for num, sig, stype, dirn in pin_defs:
        p = Pin(
            pin_number=num, signal_name=sig, signal_type=stype,
            direction=dirn, connector_id=c.id,
        )
        db_session.add(p)
        pins.append(p)
    db_session.commit()
    db_session.refresh(c)
    for p in pins:
        db_session.refresh(p)
    return c, pins


@pytest.fixture()
def test_connector_b(db_session, test_project, test_unit_b):
    """Matching connector J1 on C2P-200 with same signal names."""
    c = Connector(
        connector_id="CONN-002", designator="J1", name="1553 Bus A",
        connector_type="mil_dtl_38999_series_iii", gender="male_pin",
        total_contacts=25, unit_id=test_unit_b.id, project_id=test_project.id,
    )
    db_session.add(c)
    db_session.flush()
    pin_defs = [
        ("1", "28V_PWR", "power_primary", "power_sink"),
        ("2", "28V_RTN", "power_return", "power_return"),
        ("3", "1553A_HI", "signal_digital_differential", "bidirectional"),
        ("4", "1553A_LO", "signal_digital_differential", "bidirectional"),
        ("5", "SPARE_1", "spare", "no_connect"),
    ]
    pins = []
    for num, sig, stype, dirn in pin_defs:
        p = Pin(
            pin_number=num, signal_name=sig, signal_type=stype,
            direction=dirn, connector_id=c.id,
        )
        db_session.add(p)
        pins.append(p)
    db_session.commit()
    db_session.refresh(c)
    for p in pins:
        db_session.refresh(p)
    return c, pins


@pytest.fixture()
def test_bus(db_session, test_project, test_unit, test_connector_a):
    conn, pins = test_connector_a
    bd = BusDefinition(
        bus_def_id="BUS-001", name="1553 Bus A", protocol="mil_std_1553b",
        bus_role="remote_terminal", bus_address="RT05", data_rate="1 Mbps",
        word_size_bits=16, bus_name_network="MUX_BUS_A",
        unit_id=test_unit.id, project_id=test_project.id,
        data_rate_actual_bps=1000000,
    )
    db_session.add(bd)
    db_session.flush()
    # Assign 1553 pins to the bus
    for p in pins[2:4]:  # 1553A_HI, 1553A_LO
        pa = PinBusAssignment(pin_id=p.id, bus_def_id=bd.id, pin_role="data_positive" if "HI" in p.signal_name else "data_negative")
        db_session.add(pa)
    db_session.commit()
    db_session.refresh(bd)
    return bd


@pytest.fixture()
def test_message(db_session, test_project, test_unit, test_bus):
    msg = MessageDefinition(
        msg_def_id="MSG-001", label="Target Track", mnemonic="TGT_TRK",
        direction="transmit", subaddress=5, word_count=16, rate_hz=50.0,
        latency_max_ms=20.0, priority="mission_critical",
        scheduling="periodic_synchronous",
        bus_def_id=test_bus.id, unit_id=test_unit.id, project_id=test_project.id,
    )
    db_session.add(msg)
    db_session.flush()
    # Add 2 fields
    f1 = MessageField(
        field_name="target_id", label="Target ID", data_type="uint16",
        word_number=1, bit_offset=0, bit_length=16, message_id=msg.id,
        field_order=1, min_value=0, max_value=65534,
    )
    f2 = MessageField(
        field_name="range_m", label="Range", data_type="float32",
        word_number=2, bit_offset=0, bit_length=32, message_id=msg.id,
        field_order=2, unit_of_measure="meters", scale_factor=0.1,
        min_value=0, max_value=300000,
    )
    db_session.add_all([f1, f2])
    db_session.commit()
    db_session.refresh(msg)
    return msg


@pytest.fixture()
def test_harness(db_session, test_project, test_unit, test_unit_b, test_connector_a, test_connector_b):
    conn_a, _ = test_connector_a
    conn_b, _ = test_connector_b
    h = WireHarness(
        harness_id="HAR-001", name="RSP-100 to C2P-200",
        from_unit_id=test_unit.id, from_connector_id=conn_a.id,
        to_unit_id=test_unit_b.id, to_connector_id=conn_b.id,
        project_id=test_project.id, cable_type="MIL-DTL-27500",
        overall_length_m=2.5, overall_length_max_m=3.0,
    )
    db_session.add(h)
    db_session.commit()
    db_session.refresh(h)
    return h


# ══════════════════════════════════════════════════════════════
#  1. System CRUD (5 tests)
# ══════════════════════════════════════════════════════════════

class TestSystemCRUD:

    def test_create_system(self, client, auth_headers, test_project):
        resp = client.post(
            f"/api/v1/interfaces/systems?project_id={test_project.id}",
            json={"name": "Nav Subsystem", "system_type": "subsystem", "abbreviation": "NAV"},
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["name"] == "Nav Subsystem"
        assert data["system_id"].startswith("SYS-")
        assert data["unit_count"] == 0

    def test_list_systems(self, client, auth_headers, test_project, test_system):
        resp = client.get(
            "/api/v1/interfaces/systems",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) >= 1
        assert any(s["name"] == "Radar Subsystem" for s in data)

    def test_get_system_detail(self, client, auth_headers, test_system, test_unit):
        resp = client.get(f"/api/v1/interfaces/systems/{test_system.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "Radar Subsystem"
        assert len(data["units"]) == 1
        assert data["units"][0]["designation"] == "RSP-100"

    def test_update_system(self, client, auth_headers, test_system):
        resp = client.patch(
            f"/api/v1/interfaces/systems/{test_system.id}",
            json={"responsible_org": "Lockheed Martin"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["responsible_org"] == "Lockheed Martin"

    def test_delete_system_blocked_with_units(self, client, auth_headers, test_system, test_unit):
        resp = client.delete(
            f"/api/v1/interfaces/systems/{test_system.id}",
            headers=auth_headers,
        )
        assert resp.status_code == 409, "Should refuse deletion when units exist"


# ══════════════════════════════════════════════════════════════
#  2. Unit CRUD (8 tests)
# ══════════════════════════════════════════════════════════════

class TestUnitCRUD:

    def test_create_unit(self, client, auth_headers, test_project, test_system):
        resp = client.post(
            f"/api/v1/interfaces/units?project_id={test_project.id}",
            json={
                "name": "Antenna Assembly", "designation": "ANT-300",
                "part_number": "ANT-300-C1", "manufacturer": "Northrop",
                "unit_type": "antenna", "system_id": test_system.id,
                "mass_kg": 12.5, "power_watts_nominal": 0,
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["designation"] == "ANT-300"
        assert data["unit_id"].startswith("UNIT-")

    def test_create_duplicate_designation_rejected(self, client, auth_headers, test_project, test_system, test_unit):
        resp = client.post(
            f"/api/v1/interfaces/units?project_id={test_project.id}",
            json={
                "name": "Duplicate", "designation": "RSP-100",
                "part_number": "X", "manufacturer": "X",
                "unit_type": "processor", "system_id": test_system.id,
            },
            headers=auth_headers,
        )
        assert resp.status_code == 409, "Duplicate designation must be rejected"

    def test_list_units(self, client, auth_headers, test_project, test_unit):
        resp = client.get(
            "/api/v1/interfaces/units",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert len(resp.json()) >= 1

    def test_list_units_search(self, client, auth_headers, test_project, test_unit):
        resp = client.get(
            "/api/v1/interfaces/units",
            params={"project_id": test_project.id, "search": "radar"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert any(u["designation"] == "RSP-100" for u in data)

    def test_get_unit_detail(self, client, auth_headers, test_unit, test_connector_a):
        resp = client.get(f"/api/v1/interfaces/units/{test_unit.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["designation"] == "RSP-100"
        assert data["connector_count"] >= 1
        assert len(data["connectors"]) >= 1
        assert len(data["connectors"][0]["pins"]) == 5

    def test_update_unit(self, client, auth_headers, test_unit):
        resp = client.patch(
            f"/api/v1/interfaces/units/{test_unit.id}",
            json={"mass_kg": 3.1, "heritage": "Heritage from AN/APG-79"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["mass_kg"] == 3.1

    def test_delete_unit_preview(self, client, auth_headers, test_unit, test_connector_a):
        resp = client.delete(
            f"/api/v1/interfaces/units/{test_unit.id}",
            params={"confirm": False},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "preview"
        assert data["impact"]["connectors"] >= 1
        assert data["impact"]["pins"] >= 5

    def test_get_unit_specifications(self, client, auth_headers, test_unit):
        resp = client.get(f"/api/v1/interfaces/units/{test_unit.id}/specifications", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["unit_id"] == test_unit.id


# ══════════════════════════════════════════════════════════════
#  3. Connector and Pins (8 tests)
# ══════════════════════════════════════════════════════════════

class TestConnectorAndPins:

    def test_create_connector_with_inline_pins(self, client, auth_headers, test_unit, test_project):
        resp = client.post(
            "/api/v1/interfaces/connectors",
            json={
                "designator": "J9", "connector_type": "d_sub_9",
                "gender": "female_socket", "total_contacts": 3,
                "unit_id": test_unit.id,
                "pins": [
                    {"pin_number": "1", "signal_name": "TX", "signal_type": "serial_data", "direction": "output"},
                    {"pin_number": "2", "signal_name": "RX", "signal_type": "serial_data", "direction": "input"},
                    {"pin_number": "3", "signal_name": "GND", "signal_type": "signal_ground", "direction": "ground"},
                ],
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["pin_count"] == 3
        assert len(data["pins"]) == 3

    def test_list_connectors(self, client, auth_headers, test_connector_a):
        conn, _ = test_connector_a
        resp = client.get(
            "/api/v1/interfaces/connectors",
            params={"unit_id": conn.unit_id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert len(resp.json()) >= 1

    def test_get_connector_with_pins(self, client, auth_headers, test_connector_a):
        conn, _ = test_connector_a
        resp = client.get(f"/api/v1/interfaces/connectors/{conn.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["designator"] == "J1"
        assert len(data["pins"]) == 5

    def test_get_pinout(self, client, auth_headers, test_connector_a):
        conn, _ = test_connector_a
        resp = client.get(f"/api/v1/interfaces/connectors/{conn.id}/pinout", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_pins"] == 5
        assert data["pin_summary"]["power"] >= 1
        assert data["pin_summary"]["spare"] >= 1

    def test_batch_add_pins(self, client, auth_headers, test_connector_a):
        conn, _ = test_connector_a
        resp = client.post(
            f"/api/v1/interfaces/connectors/{conn.id}/pins",
            json={"pins": [
                {"pin_number": "6", "signal_name": "DISC_OUT_1", "signal_type": "discrete_output", "direction": "output"},
                {"pin_number": "7", "signal_name": "DISC_IN_1", "signal_type": "discrete_input", "direction": "input"},
            ]},
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        assert len(resp.json()) == 2

    def test_batch_add_duplicate_pin_rejected(self, client, auth_headers, test_connector_a):
        conn, _ = test_connector_a
        resp = client.post(
            f"/api/v1/interfaces/connectors/{conn.id}/pins",
            json={"pins": [
                {"pin_number": "1", "signal_name": "DUPE", "signal_type": "spare", "direction": "no_connect"},
            ]},
            headers=auth_headers,
        )
        assert resp.status_code == 409, "Duplicate pin number must be rejected"

    def test_auto_generate_pins(self, client, auth_headers, test_unit, test_project):
        # Create a connector with no pins
        c = Connector(
            connector_id="CONN-AUTO", designator="J99",
            connector_type="d_sub_9", gender="female_socket",
            total_contacts=9, unit_id=test_unit.id, project_id=test_project.id,
        )
        from tests.conftest import db_session  # noqa — handled by fixture chain
        # Use client to create via API instead
        create_resp = client.post(
            "/api/v1/interfaces/connectors",
            json={
                "designator": "JAUTO", "connector_type": "d_sub_9",
                "gender": "female_socket", "total_contacts": 4,
                "unit_id": test_unit.id,
            },
            headers=auth_headers,
        )
        conn_id = create_resp.json()["id"]
        resp = client.post(
            f"/api/v1/interfaces/connectors/{conn_id}/pins/auto-generate",
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert len(data) == 4
        assert all(p["signal_type"] == "spare" for p in data)

    def test_pin_search_across_units(self, client, auth_headers, test_project, test_connector_a, test_connector_b):
        resp = client.get(
            "/api/v1/interfaces/pins/search",
            params={"project_id": test_project.id, "signal_name": "1553A"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert len(data) >= 2, "Should find 1553A pins on both units"


# ══════════════════════════════════════════════════════════════
#  4. Bus Definition (5 tests)
# ══════════════════════════════════════════════════════════════

class TestBusDefinition:

    def test_create_bus(self, client, auth_headers, test_unit, test_project):
        resp = client.post(
            "/api/v1/interfaces/buses",
            json={
                "name": "RS-422 Link", "protocol": "rs422",
                "bus_role": "master", "unit_id": test_unit.id,
                "data_rate": "1 Mbps",
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        assert resp.json()["bus_def_id"].startswith("BUS-")

    def test_get_bus_with_messages(self, client, auth_headers, test_bus, test_message):
        resp = client.get(f"/api/v1/interfaces/buses/{test_bus.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["message_count"] >= 1
        assert len(data["messages"]) >= 1
        assert data["messages"][0]["label"] == "Target Track"

    def test_bus_pin_assignments(self, client, auth_headers, test_bus):
        resp = client.get(f"/api/v1/interfaces/buses/{test_bus.id}", headers=auth_headers)
        data = resp.json()
        assert data["pin_assignment_count"] >= 2

    def test_bus_utilization(self, client, auth_headers, test_bus, test_message):
        resp = client.get(f"/api/v1/interfaces/buses/{test_bus.id}/utilization", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["used_bps"] > 0
        assert data["message_count"] >= 1

    def test_delete_bus_preview(self, client, auth_headers, test_bus, test_message):
        resp = client.delete(
            f"/api/v1/interfaces/buses/{test_bus.id}",
            params={"confirm": False},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["status"] == "preview"
        assert data["impact"]["messages"] >= 1


# ══════════════════════════════════════════════════════════════
#  5. Message Definition (5 tests)
# ══════════════════════════════════════════════════════════════

class TestMessageDefinition:

    def test_create_message_with_fields(self, client, auth_headers, test_bus, test_unit):
        resp = client.post(
            "/api/v1/interfaces/messages",
            json={
                "label": "Status Word", "mnemonic": "STS",
                "direction": "transmit", "bus_def_id": test_bus.id,
                "unit_id": test_unit.id, "word_count": 1, "rate_hz": 100.0,
                "fields": [
                    {"field_name": "mode", "data_type": "enum_coded", "bit_length": 4, "message_id": 0},
                    {"field_name": "health", "data_type": "uint8", "bit_length": 8, "message_id": 0},
                ],
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["field_count"] == 2
        assert data["total_bits"] == 12

    def test_get_message_with_fields(self, client, auth_headers, test_message):
        resp = client.get(f"/api/v1/interfaces/messages/{test_message.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["label"] == "Target Track"
        assert len(data["fields"]) == 2

    def test_list_messages_by_bus(self, client, auth_headers, test_bus, test_message):
        resp = client.get(
            "/api/v1/interfaces/messages",
            params={"bus_id": test_bus.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert len(resp.json()) >= 1

    def test_byte_map(self, client, auth_headers, test_message):
        resp = client.get(f"/api/v1/interfaces/messages/{test_message.id}/byte-map", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["total_fields"] == 2
        assert data["total_bits_used"] == 48  # 16 + 32
        assert len(data["layout"]) >= 1

    def test_batch_add_fields(self, client, auth_headers, test_message):
        resp = client.post(
            f"/api/v1/interfaces/messages/{test_message.id}/fields",
            json={"fields": [
                {"field_name": "spare_w5", "data_type": "spare", "bit_length": 16, "message_id": test_message.id, "is_spare": True},
            ]},
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        assert len(resp.json()) == 1


# ══════════════════════════════════════════════════════════════
#  6. Wire Harness (6 tests)
# ══════════════════════════════════════════════════════════════

class TestWireHarness:

    def test_create_harness(self, client, auth_headers, test_unit, test_unit_b, test_connector_a, test_connector_b):
        conn_a, _ = test_connector_a
        conn_b, _ = test_connector_b
        resp = client.post(
            "/api/v1/interfaces/harnesses",
            json={
                "name": "Test Harness", "from_unit_id": test_unit.id,
                "from_connector_id": conn_a.id, "to_unit_id": test_unit_b.id,
                "to_connector_id": conn_b.id,
            },
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["harness_id"].startswith("HAR-")
        assert data["wire_count"] == 0

    def test_create_harness_wrong_connector_rejected(self, client, auth_headers, test_unit, test_unit_b, test_connector_a, test_connector_b):
        conn_a, _ = test_connector_a
        conn_b, _ = test_connector_b
        resp = client.post(
            "/api/v1/interfaces/harnesses",
            json={
                "name": "Bad Harness", "from_unit_id": test_unit.id,
                "from_connector_id": conn_b.id,  # Wrong — conn_b belongs to unit_b
                "to_unit_id": test_unit_b.id,
                "to_connector_id": conn_a.id,     # Wrong
            },
            headers=auth_headers,
        )
        assert resp.status_code == 400, "Should reject mismatched connector/unit"

    def test_get_harness_detail(self, client, auth_headers, test_harness):
        resp = client.get(f"/api/v1/interfaces/harnesses/{test_harness.id}", headers=auth_headers)
        assert resp.status_code == 200
        data = resp.json()
        assert data["name"] == "RSP-100 to C2P-200"

    def test_batch_create_wires(self, client, auth_headers, test_harness, test_connector_a, test_connector_b):
        _, pins_a = test_connector_a
        _, pins_b = test_connector_b
        resp = client.post(
            f"/api/v1/interfaces/harnesses/{test_harness.id}/wires",
            json={"wires": [
                {"wire_number": "W001", "signal_name": "28V_PWR", "wire_type": "power_positive",
                 "from_pin_id": pins_a[0].id, "to_pin_id": pins_b[0].id},
                {"wire_number": "W002", "signal_name": "28V_RTN", "wire_type": "power_return",
                 "from_pin_id": pins_a[1].id, "to_pin_id": pins_b[1].id},
            ]},
            headers=auth_headers,
        )
        assert resp.status_code == 201, resp.text
        data = resp.json()
        assert data["count"] == 2

    def test_wire_wrong_connector_rejected(self, client, auth_headers, test_harness, test_connector_a, test_connector_b):
        _, pins_a = test_connector_a
        _, pins_b = test_connector_b
        resp = client.post(
            f"/api/v1/interfaces/harnesses/{test_harness.id}/wires",
            json={"wires": [
                {"wire_number": "WBAD", "signal_name": "BAD", "wire_type": "signal_single",
                 "from_pin_id": pins_b[0].id,  # Wrong — from should be on from_connector
                 "to_pin_id": pins_a[0].id},
            ]},
            headers=auth_headers,
        )
        assert resp.status_code == 400, "Wire with wrong connector pins must be rejected"

    def test_delete_harness_preview(self, client, auth_headers, test_harness):
        resp = client.delete(
            f"/api/v1/interfaces/harnesses/{test_harness.id}",
            params={"confirm": False},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["status"] == "preview"


# ══════════════════════════════════════════════════════════════
#  7. Auto-Wire (3 tests)
# ══════════════════════════════════════════════════════════════

class TestAutoWire:

    def test_auto_wire_matches_by_signal_name(self, client, auth_headers, test_harness):
        resp = client.post(
            f"/api/v1/interfaces/harnesses/{test_harness.id}/auto-wire",
            headers=auth_headers,
        )
        assert resp.status_code == 200, resp.text
        data = resp.json()
        # Should match: 28V_PWR, 28V_RTN, 1553A_HI, 1553A_LO (4 matches)
        # SPARE_1 should NOT match (skipped)
        assert data["matched"] >= 3, f"Expected >= 3 matches, got {data['matched']}"
        assert len(data["wires_created"]) >= 3

    def test_auto_wire_skips_spares(self, client, auth_headers, test_harness):
        resp = client.post(
            f"/api/v1/interfaces/harnesses/{test_harness.id}/auto-wire",
            headers=auth_headers,
        )
        data = resp.json()
        wire_signals = [w["signal_name"] for w in data["wires_created"]]
        assert "SPARE_1" not in wire_signals, "Spare pins should not be auto-wired"

    def test_auto_wire_reports_unmatched(self, client, auth_headers, test_harness):
        resp = client.post(
            f"/api/v1/interfaces/harnesses/{test_harness.id}/auto-wire",
            headers=auth_headers,
        )
        data = resp.json()
        # unmatched_from should contain SPARE_1
        unmatched_signals = [p["signal_name"] for p in data.get("unmatched_from", [])]
        assert "SPARE_1" in unmatched_signals


# ══════════════════════════════════════════════════════════════
#  8. Auto-Requirement Generation (8 tests) — CRITICAL
# ══════════════════════════════════════════════════════════════

class TestAutoRequirementGeneration:

    def _create_wires_and_get_reqs(self, db_session, test_harness, test_connector_a, test_connector_b, test_user, test_bus=None):
        """Helper: create wires and run the auto-requirement generator."""
        from app.services.interface.auto_requirements import AutoRequirementGenerator
        _, pins_a = test_connector_a
        _, pins_b = test_connector_b

        wires = []
        for i, (pa, pb) in enumerate(zip(pins_a[:4], pins_b[:4])):
            w = Wire(
                wire_number=f"W{i+1:03d}", signal_name=pa.signal_name,
                wire_type="power_positive" if "PWR" in pa.signal_name else
                          "power_return" if "RTN" in pa.signal_name else
                          "signal_twisted_pair_a",
                from_pin_id=pa.id, to_pin_id=pb.id, harness_id=test_harness.id,
            )
            db_session.add(w)
            wires.append(w)
        db_session.flush()

        gen = AutoRequirementGenerator(db_session, test_harness.project_id, test_user)
        result = gen.on_wires_created(test_harness, wires)
        db_session.commit()
        return result

    def test_wire_creates_harness_requirement(self, db_session, test_user, test_harness,
                                               test_connector_a, test_connector_b, test_bus):
        result = self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user, test_bus
        )
        assert result["requirements_generated"] >= 1, "Should generate at least harness-level req"
        titles = [r["title"] for r in result["requirements"]]
        assert any("Harness" in t or "HAR" in t for t in titles), f"Expected harness req, got: {titles}"

    def test_wire_creates_bus_requirement(self, db_session, test_user, test_harness,
                                          test_connector_a, test_connector_b, test_bus):
        result = self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user, test_bus
        )
        titles = [r["title"] for r in result["requirements"]]
        assert any("1553" in t or "Data Link" in t for t in titles), f"Expected bus connection req, got: {titles}"

    def test_power_wire_creates_requirement(self, db_session, test_user, test_harness,
                                             test_connector_a, test_connector_b):
        result = self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user
        )
        titles = [r["title"] for r in result["requirements"]]
        assert any("Power" in t for t in titles), f"Expected power wire req, got: {titles}"

    def test_auto_reqs_have_quality_scores(self, db_session, test_user, test_harness,
                                            test_connector_a, test_connector_b):
        result = self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user
        )
        for r in result["requirements"]:
            assert r["quality_score"] > 0, f"Auto-req '{r['title']}' has quality_score=0"

    def test_auto_reqs_linked_to_entities(self, db_session, test_user, test_harness,
                                           test_connector_a, test_connector_b):
        self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user
        )
        links = db_session.query(InterfaceRequirementLink).filter(
            InterfaceRequirementLink.auto_generated.is_(True),
        ).all()
        assert len(links) >= 1, "Should create InterfaceRequirementLink records"
        assert all(lk.status == "pending_review" for lk in links)

    def test_auto_reqs_create_verifications(self, db_session, test_user, test_harness,
                                             test_connector_a, test_connector_b):
        result = self._create_wires_and_get_reqs(
            db_session, test_harness, test_connector_a, test_connector_b, test_user
        )
        assert result["verifications_generated"] >= 1, "Should create verification records"
        verifs = db_session.query(Verification).filter(
            Verification.status == "planned"
        ).all()
        assert len(verifs) >= 1

    def test_unit_specs_create_env_requirements(self, db_session, test_user, test_unit):
        from app.services.interface.auto_requirements import AutoRequirementGenerator
        gen = AutoRequirementGenerator(db_session, test_unit.project_id, test_user)
        result = gen.on_unit_created(test_unit)
        db_session.commit()
        assert result["total_generated"] >= 3, (
            f"Expected >= 3 env/EMI reqs from unit specs, got {result['total_generated']}"
        )
        # Should include temp, vibration, and at least one EMI
        titles = [r["title"] for r in result["requirements"]]
        assert any("Temperature" in t for t in titles), f"Expected temp req, got: {titles}"

    def test_unit_emi_creates_emi_requirements(self, db_session, test_user, test_unit):
        from app.services.interface.auto_requirements import AutoRequirementGenerator
        gen = AutoRequirementGenerator(db_session, test_unit.project_id, test_user)
        result = gen.on_unit_created(test_unit)
        db_session.commit()
        titles = [r["title"] for r in result["requirements"]]
        assert any("CE102" in t or "Emi" in t or "EMI" in t.upper() or "Conducted" in t for t in titles), (
            f"Expected EMI req from ce102 spec, got: {titles}"
        )


# ══════════════════════════════════════════════════════════════
#  9. Excel Import/Export (5 tests)
# ══════════════════════════════════════════════════════════════

class TestExcelImportExport:

    def test_download_template(self, client, auth_headers):
        resp = client.post("/api/v1/interfaces/io/import/template", headers=auth_headers)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")
        assert len(resp.content) > 1000, "Template should be a real xlsx file"

    def test_template_has_four_sheets(self, client, auth_headers):
        from openpyxl import load_workbook
        resp = client.post("/api/v1/interfaces/io/import/template", headers=auth_headers)
        wb = load_workbook(filename=io.BytesIO(resp.content))
        assert "Units" in wb.sheetnames
        assert "Connectors" in wb.sheetnames
        assert "Buses" in wb.sheetnames
        assert "Messages" in wb.sheetnames
        wb.close()

    def test_export_units(self, client, auth_headers, test_project, test_unit, test_connector_a):
        resp = client.get(
            "/api/v1/interfaces/io/export/units",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    def test_export_units_has_data(self, client, auth_headers, test_project, test_unit, test_connector_a):
        from openpyxl import load_workbook
        resp = client.get(
            "/api/v1/interfaces/io/export/units",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        wb = load_workbook(filename=io.BytesIO(resp.content))
        ws = wb["Units"]
        assert ws.max_row >= 2, "Units sheet should have at least one data row"
        # Verify the unit designation is in the data
        found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1] == "RSP-100":
                found = True
                break
        assert found, "RSP-100 should appear in exported Units sheet"
        wb.close()

    def test_export_icd_data(self, client, auth_headers, test_project, test_unit, test_connector_a, test_bus, test_message):
        resp = client.get(
            "/api/v1/interfaces/io/export/icd-data",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(resp.content))
        assert len(wb.sheetnames) >= 8, f"ICD export should have >= 8 sheets, got {wb.sheetnames}"
        wb.close()


# ══════════════════════════════════════════════════════════════
#  10. Signal Trace (3 tests)
# ══════════════════════════════════════════════════════════════

class TestSignalTrace:

    def _setup_wired_harness(self, db_session, test_harness, test_connector_a, test_connector_b):
        """Create wires for signal trace testing."""
        _, pins_a = test_connector_a
        _, pins_b = test_connector_b
        for i, (pa, pb) in enumerate(zip(pins_a[:4], pins_b[:4])):
            w = Wire(
                wire_number=f"WT{i+1:03d}", signal_name=pa.signal_name,
                wire_type="signal_single", from_pin_id=pa.id,
                to_pin_id=pb.id, harness_id=test_harness.id,
            )
            db_session.add(w)
        db_session.commit()

    def test_signal_trace_finds_path(self, client, auth_headers, test_project,
                                     test_harness, test_connector_a, test_connector_b, db_session):
        self._setup_wired_harness(db_session, test_harness, test_connector_a, test_connector_b)
        resp = client.get(
            "/api/v1/interfaces/signal-trace",
            params={"project_id": test_project.id, "signal_name": "1553A_HI"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["signal_name"] == "1553A_HI"
        assert len(data["path"]) >= 1, "Should find at least one hop in the trace"

    def test_signal_trace_no_match(self, client, auth_headers, test_project):
        resp = client.get(
            "/api/v1/interfaces/signal-trace",
            params={"project_id": test_project.id, "signal_name": "NONEXISTENT_SIGNAL"},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert len(resp.json()["path"]) == 0

    def test_coverage_endpoint(self, client, auth_headers, test_project):
        resp = client.get(
            "/api/v1/interfaces/coverage",
            params={"project_id": test_project.id},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert "total_interfaces" in data
        assert "coverage_pct" in data
        assert isinstance(data["coverage_pct"], (int, float))
